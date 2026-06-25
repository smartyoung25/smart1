import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

wb = Workbook()

# ── 색상 팔레트 ────────────────────────────────────────────────
GREEN_D  = '1F6B3B'   # 진녹색 (헤더)
GREEN_M  = '2D8653'   # 중녹색
GREEN_L  = 'D6F0E0'   # 연녹색 (소계 행)
ORANGE_D = 'C25B00'   # 비용 헤더
ORANGE_L = 'FFF0E0'   # 비용 소계 행
YELLOW   = 'FFF9C4'   # 입력셀
BLUE_D   = '0D47A1'   # KPI 헤더
BLUE_L   = 'E3F2FD'   # KPI 셀
GRAY_H   = 'F5F5F5'   # 그레이 행
WHITE    = 'FFFFFF'
RED_L    = 'FFE0E0'

# ── 스타일 유틸 ────────────────────────────────────────────────
def bd(clr='BBBBBB'):
    s = Side(border_style='thin', color=clr)
    return Border(left=s, right=s, top=s, bottom=s)
def hd_fill(hex_clr): return PatternFill('solid', fgColor=hex_clr)
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def left(): return Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
def right(): return Alignment(horizontal='right', vertical='center')
def fmt_won(ws, row, col):
    ws.cell(row=row, column=col).number_format = '#,##0'
def fmt_pct(ws, row, col):
    ws.cell(row=row, column=col).number_format = '0.0%'

def style_cell(cell, bold=False, italic=False, size=10, color='000000',
               bg=None, align=None, fmt=None, border=True):
    cell.font = Font(name='맑은 고딕', bold=bold, italic=italic,
                     size=size, color=color)
    if bg: cell.fill = hd_fill(bg)
    if align: cell.alignment = align
    if fmt:   cell.number_format = fmt
    if border: cell.border = bd()

# ── 헬퍼: 셀 쓰기 + 스타일 한 번에 ───────────────────────────
def wc(ws, row, col, val, bold=False, italic=False, size=10, color='000000',
       bg=None, align=None, fmt=None, border=True):
    c = ws.cell(row=row, column=col, value=val)
    style_cell(c, bold, italic, size, color, bg, align, fmt, border)
    return c

def merge_wc(ws, r1, c1, r2, c2, val, **kw):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    wc(ws, r1, c1, val, align=center(), **kw)

# ═══════════════════════════════════════════════════════════════
# 시트 1: 경영성과 요약 (Summary)
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = '경영성과 요약'
ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 16
ws1.column_dimensions['C'].width = 16
ws1.column_dimensions['D'].width = 16
ws1.column_dimensions['E'].width = 14
ws1.column_dimensions['F'].width = 14
ws1.freeze_panes = 'A4'

# 타이틀
ws1.row_dimensions[1].height = 36
ws1.row_dimensions[2].height = 18
merge_wc(ws1,1,1,1,6,'🌱 스마트팜 경영성과 체계화 분석표 — 방울토마토 장호림',
         bold=True, size=14, color=WHITE, bg=GREEN_D, border=True)
merge_wc(ws1,2,1,2,6,
         '농가: 홍길동 | 재배면적: 5,300㎡(53a) | 조사면적: 900㎡(9a) | 품종: 방울토마토(촉성)',
         italic=True, size=9, color='444444', bg='F0FFF4', border=True)

# ── A. KPI 카드 행 ─────────────────────────────────────────────
ws1.row_dimensions[3].height = 16
ws1.row_dimensions[4].height = 40
ws1.row_dimensions[5].height = 28
KPI_HEADS = ['총수입(원)', '경영비(원)', '소득(원)', '소득률(%)', '순수익률(%)']
KPI_VALS  = [537144230, 344293421, 192850809, 0.3590, 0.2280]
KPI_BG    = [GREEN_L, ORANGE_L, BLUE_L, BLUE_L, BLUE_L]
wc(ws1,3,1,'[ KPI 핵심지표 ]',bold=True,size=9,color=GREEN_D,bg='F0FFF4',border=False)
for i,(h,v,bg) in enumerate(zip(KPI_HEADS, KPI_VALS, KPI_BG)):
    col = i+2
    wc(ws1,3,col,h,bold=True,size=8,color='444444',bg=GRAY_H,align=center())
    c = wc(ws1,4,col,v,bold=True,size=13,align=center(),bg=bg,
           fmt='#,##0' if i<3 else '0.0%')
    if i==2: c.font = Font(name='맑은 고딕',bold=True,size=13,color=GREEN_D)
    if i==3 or i==4: c.font = Font(name='맑은 고딕',bold=True,size=13,color=BLUE_D)

# 10a 당 KPI
wc(ws1,5,1,'(10a 단위면적 기준)',italic=True,size=8,color='888888',border=False)
KPI10A = ['100,714,543', '64,555,016', '36,159,527', '—', '—']
for i,v in enumerate(KPI10A):
    wc(ws1,5,i+2,v,italic=True,size=8,color='666666',align=center(),bg=GRAY_H)

# ── B. 매출 구조 ───────────────────────────────────────────────
R = 7
ws1.row_dimensions[R].height = 20
merge_wc(ws1,R,1,R,6,'▶ 매출(수입) 구조',bold=True,size=11,color=WHITE,bg=GREEN_M)
R += 1
heads = ['항목','구분','수량','단가(원)','금액(원)','구성비(%)']
for i,h in enumerate(heads):
    wc(ws1,R,i+1,h,bold=True,size=9,color=WHITE,bg=GREEN_D,align=center())
R+=1
revenue_rows = [
    ('총수입','주산물 — 방울토마토','92,176 kg','5,827원/kg',537144230,1.0),
    ('','부산물','—','—',0,0.0),
    ('합계','','','',537144230,1.0),
]
for row in revenue_rows:
    is_total = row[0]=='합계'
    bg = GREEN_L if is_total else WHITE
    bold = is_total
    for j,v in enumerate(row):
        fmt = '#,##0' if j==4 else ('0.0%' if j==5 else None)
        wc(ws1,R,j+1,v if not (j==4 and v==0) else '—',
           bold=bold,bg=bg,align=right() if j>=4 else left(),fmt=fmt)
    R+=1

# ── C. 비용 구조 ───────────────────────────────────────────────
R+=1
ws1.row_dimensions[R].height = 20
merge_wc(ws1,R,1,R,6,'▶ 비용 구조 (경영비 + 자가노동·자본용역)',bold=True,size=11,color=WHITE,bg=ORANGE_D)
R+=1
for i,h in enumerate(heads):
    wc(ws1,R,i+1,h,bold=True,size=9,color=WHITE,bg=ORANGE_D,align=center())
R+=1

cost_data = [
    # (대분류, 세목, 금액, 비고)
    ('Ⅰ 중간재비','종자·종묘비',17940000,'방울토마토 모종 19,500주'),
    ('','보통비료(무기질)',15655500,'복합비료·요소 등'),
    ('','부산물비료(유기질)',0,'—'),
    ('','농약비',4748500,'살충제+살균제(친환경인증 포함)'),
    ('','수도광열비',41410310,'전기·경유·가스 등'),
    ('','기타재료비',50121104,'하우스비닐·보온시설 등'),
    ('','소농구비',1000000,'—'),
    ('','대농구상각비',19685000,'방제기·자동차 등'),
    ('','영농시설상각비',50068632,'하우스·베드·관수·양액·환경제어'),
    ('','수리·유지비',0,'—'),
    ('','기타비용',13030000,'—'),
    ('  소계','',213659046,''),
    ('Ⅱ 위탁영농비·임차료','임차료',0,'—'),
    ('','위탁영농비',0,'—'),
    ('  소계','',0,''),
    ('Ⅲ 고용노동비','고용노임',130634375,'남 7,280h × 10,625원 + 여 5,015h × 10,625원'),
    ('  소계','',130634375,''),
    ('━ 경영비 합계','',344293421,''),
    ('Ⅳ 자가노동비','자가노임',31816944,'남 1,368h × 23,258원'),
    ('Ⅴ 자본용역비','유동자본용역비',6863495,''),
    ('','고정자본용역비',30180748,''),
    ('','토지자본용역비',1500000,''),
    ('  소계','',38544243,''),
    ('━ 생산비 합계','',414654607,''),
]

TOTAL = 537144230
for row in cost_data:
    大, 세목, 금액, 비고 = row
    is_subtotal = '소계' in 大
    is_total = '합계' in 大
    bg = ORANGE_L if is_subtotal else (RED_L if is_total else WHITE)
    bold = is_subtotal or is_total
    ratio = 금액/TOTAL if 금액 else None
    ws1.cell(row=R,column=1).value = 大
    ws1.cell(row=R,column=2).value = 세목
    ws1.cell(row=R,column=3).value = '—'
    ws1.cell(row=R,column=4).value = '—'
    ws1.cell(row=R,column=5).value = 금액 if 금액 else '—'
    ws1.cell(row=R,column=6).value = ratio
    ws1.cell(row=R,column=7,value=비고) if False else None
    for c in range(1,7):
        cell = ws1.cell(row=R,column=c)
        cell.font = Font(name='맑은 고딕',bold=bold,size=9,
                         color=ORANGE_D if is_total else '000000')
        cell.fill = hd_fill(bg)
        cell.border = bd()
        if c==5 and isinstance(cell.value, int):
            cell.number_format = '#,##0'
            cell.alignment = right()
        elif c==6 and isinstance(cell.value, float):
            cell.number_format = '0.0%'
            cell.alignment = right()
        else:
            cell.alignment = left() if c<=2 else right()
    R+=1

# ── D. 소득 지표 ───────────────────────────────────────────────
R+=1
merge_wc(ws1,R,1,R,6,'▶ 소득·수익 지표',bold=True,size=11,color=WHITE,bg=BLUE_D)
R+=1
ind_heads = ['지표','산식','금액(원)','단위면적 10a','비율(%)','해설']
for i,h in enumerate(ind_heads):
    wc(ws1,R,i+1,h,bold=True,size=9,color=WHITE,bg=BLUE_D,align=center())
R+=1
indicators = [
    ('소득','총수입 - 경영비',192850809,36159527,0.3590,'경영성과 핵심지표'),
    ('부가가치','소득 + 고용노동비',323485184,60653472,0.6022,'노동과 자본의 창출가치'),
    ('순수익','총수입 - 생산비',122489623,22966804,0.2280,'기회비용 차감 후 순이익'),
    ('경영비율','경영비 / 총수입',344293421,64555016,0.6410,'경영비 부담률'),
]
for ind in indicators:
    지표,산식,금액,per10a,ratio,해설 = ind
    wc(ws1,R,1,지표,bold=True,size=10,bg=BLUE_L)
    wc(ws1,R,2,산식,size=9,italic=True,color='444444',bg=WHITE)
    c=ws1.cell(row=R,column=3,value=금액); c.number_format='#,##0'; c.alignment=right(); c.font=Font(name='맑은 고딕',bold=True,size=10,color=BLUE_D); c.border=bd(); c.fill=hd_fill(BLUE_L)
    c=ws1.cell(row=R,column=4,value=per10a); c.number_format='#,##0'; c.alignment=right(); c.font=Font(name='맑은 고딕',size=9); c.border=bd()
    c=ws1.cell(row=R,column=5,value=ratio); c.number_format='0.0%'; c.alignment=center(); c.font=Font(name='맑은 고딕',bold=True,size=10); c.border=bd()
    wc(ws1,R,6,해설,size=9,color='555555',align=left())
    R+=1

# ── 자동필터 해제, 인쇄 설정 ──────────────────────────────────
ws1.print_title_rows = '1:3'
ws1.page_setup.fitToPage = True
ws1.page_setup.fitToWidth = 1

# ═══════════════════════════════════════════════════════════════
# 시트 2: 비용 세부 구조
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('비용 세부 구조')
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 26
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 15
ws2.column_dimensions['F'].width = 18
ws2.column_dimensions['G'].width = 12
ws2.freeze_panes = 'A3'

merge_wc(ws2,1,1,1,7,'비용 세부 구조 — 방울토마토 장호림 (53a 기준)',
         bold=True,size=13,color=WHITE,bg=ORANGE_D,border=True)

hs = ['대분류','비목','금액(원)','10a 단위(원)','비중(%)','세부내용','비고']
for i,h in enumerate(hs):
    wc(ws2,2,i+1,h,bold=True,size=9,color=WHITE,bg=ORANGE_D,align=center())

detail_rows = [
    # 중간재비
    ('중간재비','종자·종묘비',17940000,3363750,0.0334,'방울토마토 모종 19,500주','주당 920원'),
    ('','보통비료(무기질)',15655500,2935406,0.0291,'복합비료·요소·황산칼리 등','양액 혼합용'),
    ('','부산물비료(유기질)',0,0,0.0,'—','미사용'),
    ('','농약비',4748500,890344,0.0088,'살충제(친환경)+살균제','친환경인증'),
    ('','수도광열비',41410310,7764433,0.0771,'전기·경유·가스·수도료','난방비 포함'),
    ('','기타재료비',50121104,9397707,0.0933,'하우스비닐·보온시설·포장재 등','교체주기 반영'),
    ('','소농구비',1000000,187500,0.0019,'소농기구 구입·수리',''),
    ('','대농구상각비',19685000,3690938,0.0366,'방제기(80M)·자동차·지게차·선별기','감가상각 연간'),
    ('','영농시설상각비',50068632,9387869,0.0932,'하우스435M·베드125M·관수45M·양액38M·환경제어20M·전기승압32M','15~10년 내용연수'),
    ('','수리·유지비',0,0,0.0,'—',''),
    ('','기타비용',13030000,2443125,0.0243,'기타운영비·출하관련비',''),
    ('중간재비 소계','',213659046,40061071,0.3977,'',''),
    # 노동비
    ('고용노동비','남자 고용노임',77371875,14506617,0.1441,'7,280h × 10,625원/일','수확·포장·관리'),
    ('','여자 고용노임',53262500,9986719,0.0992,'5,015h × 10,625원/일','수확·선별'),
    ('고용노동비 소계','',130634375,24493945,0.2433,'',''),
    ('자가노동비','남자 자가노임',31816944,5965677,0.0593,'1,368h × 23,258원/일','경영주'),
    ('자가노동비 소계','',31816944,5965677,0.0593,'',''),
    # 임차·위탁
    ('임차·위탁','토지 임차료',0,0,0.0,'—','자가토지'),
    ('','대농기구 임차료',0,0,0.0,'—',''),
    ('','위탁영농비',0,0,0.0,'—',''),
    # 자본용역비
    ('자본용역비','유동자본용역비',6863495,1286905,0.0128,'유동자본×5%×후산출계수×재포기간',''),
    ('','고정자본용역비',30180748,5658890,0.0562,'부분현재가×5%','하우스·양액·제어'),
    ('','토지자본용역비',1500000,281250,0.0028,'자가토지 5,280㎡×기준가',''),
    ('자본용역비 소계','',38544243,7227045,0.0718,'',''),
    # 합계
    ('경영비 합계','',344293421,64555016,0.6410,'',''),
    ('생산비 합계','',414654607,77747739,0.7723,'자가노동+자본용역 포함',''),
]

for i, row in enumerate(detail_rows):
    R3 = i+3
    대분류,비목,금액,per10a,비중,세부,비고 = row
    is_sub = '소계' in 대분류
    is_total = '합계' in 대분류
    bg = ORANGE_L if is_sub else (RED_L if is_total else WHITE)
    bold = is_sub or is_total
    vals = [대분류, 비목, 금액, per10a, 비중, 세부, 비고]
    fmts = [None, None, '#,##0', '#,##0', '0.0%', None, None]
    aligns = [left(),left(),right(),right(),right(),left(),left()]
    for c,(v,f,a) in enumerate(zip(vals,fmts,aligns)):
        cell = ws2.cell(row=R3, column=c+1, value=v if v!=0 else '—' if f=='#,##0' else v)
        cell.font = Font(name='맑은 고딕', bold=bold, size=9,
                         color=ORANGE_D if is_total else '000000')
        cell.fill = hd_fill(bg)
        cell.border = bd()
        if f and isinstance(cell.value,(int,float)):
            cell.number_format = f
        cell.alignment = a

# ═══════════════════════════════════════════════════════════════
# 시트 3: 매출 구조
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('매출 구조')
ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 24
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 14
ws3.column_dimensions['E'].width = 14
ws3.column_dimensions['F'].width = 12
ws3.column_dimensions['G'].width = 18
ws3.freeze_panes = 'A3'

merge_wc(ws3,1,1,1,7,'매출 구조 — 방울토마토 장호림 (53a 기준)',
         bold=True,size=13,color=WHITE,bg=GREEN_D,border=True)
rev_heads = ['구분','품목','생산량','판매단가(원)','매출액(원)','구성비(%)','비고']
for i,h in enumerate(rev_heads):
    wc(ws3,2,i+1,h,bold=True,size=9,color=WHITE,bg=GREEN_D,align=center())

rev_rows = [
    ('주산물','방울토마토','92,176 kg','5,827원/kg',537144230,1.0,'촉성재배, 상품화율 100%'),
    ('','(단위면적 10a)','17,283 kg','5,827원/kg',100714543,'—',''),
    ('부산물','—','—','—',0,0.0,'미발생'),
    ('출하방법별','도매시장','','','','','주요 출하처'),
    ('','직거래·마트','','','','',''),
    ('','농협계통','','','','',''),
    ('총수입','',537144230,'','',1.0,''),
]
for i,row in enumerate(rev_rows):
    R3 = i+3
    is_total = row[0]=='총수입'
    bg = GREEN_L if is_total else (GRAY_H if row[0]=='출하방법별' else WHITE)
    bold = is_total
    for c,v in enumerate(row):
        cell = ws3.cell(row=R3, column=c+1, value=v if v!=0 else '—')
        cell.font = Font(name='맑은 고딕', bold=bold, size=9, color=GREEN_D if is_total else '000000')
        cell.fill = hd_fill(bg)
        cell.border = bd()
        if c in [4] and isinstance(v,(int,float)):
            cell.number_format = '#,##0'
            cell.alignment = right()
        elif c==5 and isinstance(v,float):
            cell.number_format = '0.0%'
            cell.alignment = right()
        else:
            cell.alignment = left() if c<=1 else right()

# 출하 구성 입력란 안내
R3 = len(rev_rows)+4
merge_wc(ws3,R3,1,R3,7,'▶ 출하처별 매출 구성 (직접 입력)',bold=True,size=10,color=WHITE,bg=GREEN_M,border=True)
R3+=1
out_heads=['출하처','출하 비율(%)','출하금액(원)','주요 품목','결제조건','수수료율(%)','비고']
for i,h in enumerate(out_heads):
    wc(ws3,R3,i+1,h,bold=True,size=9,color=WHITE,bg=GREEN_D,align=center())
R3+=1
channels=['도매시장(가락 등)','직거래(소비자)','농협계통','마트·B2B','포전거래','기타']
for ch in channels:
    wc(ws3,R3,1,ch,size=9,bg=YELLOW,align=left())
    for c in range(2,8):
        cell=ws3.cell(row=R3,column=c,value='')
        cell.fill=hd_fill(YELLOW)
        cell.border=bd()
        if c==2: cell.number_format='0.0%'
        if c==3: cell.number_format='#,##0'
    R3+=1

# ═══════════════════════════════════════════════════════════════
# 시트 4: 손익 분기 & 벤치마크
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('손익분기·벤치마크')
ws4.column_dimensions['A'].width = 22
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 18
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 18
ws4.freeze_panes = 'A4'

merge_wc(ws4,1,1,1,5,'손익분기·벤치마크 분석',bold=True,size=13,color=WHITE,bg=BLUE_D,border=True)
merge_wc(ws4,2,1,2,5,'경영비 기준 BEP 단가와 주요 KPI 비교 (10a 기준)',italic=True,size=9,color='444444',bg=BLUE_L,border=True)

# BEP 분석
bep_heads=['지표','산식','금액/수치','목표(개선시)','비고']
for i,h in enumerate(bep_heads):
    wc(ws4,3,i+1,h,bold=True,size=9,color=WHITE,bg=BLUE_D,align=center())

bep_rows=[
    ('재배면적(10a)','','53 a','',''),
    ('총수입','','537,144,230원','',''),
    ('경영비','','344,293,421원','',''),
    ('생산비합계','자가노동+자본용역포함','414,654,607원','',''),
    ('소득','총수입 - 경영비','192,850,809원','','소득률 35.9%'),
    ('순수익','총수입 - 생산비','122,489,623원','','순수익률 22.8%'),
    ('━ BEP 단가(경영비)','경영비÷생산량','3,735원/kg','','실제단가 5,827원'),
    ('━ BEP 단가(생산비)','생산비÷생산량','4,499원/kg','','안전마진 1,328원/kg'),
    ('BEP 생산량(경영비)','경영비÷판매단가','59,091 kg','','실생산량 92,176kg'),
    ('고용노동비 비중','고용노동비/경영비','37.9%','30%↓','인력 효율화 과제'),
    ('시설 상각 비중','상각/경영비','20.2%','','장기 감소'),
    ('노동시간(고용)','남+여 합계','12,295 h','','자동화로 절감 가능'),
]
for i,row in enumerate(bep_rows):
    R4 = i+4
    is_bep = '━ BEP' in row[0]
    bg = BLUE_L if is_bep else (GRAY_H if i%2==0 else WHITE)
    bold = is_bep
    for c,v in enumerate(row):
        cell=ws4.cell(row=R4,column=c+1,value=v)
        cell.font=Font(name='맑은 고딕',bold=bold,size=9,color=BLUE_D if is_bep else '000000')
        cell.fill=hd_fill(bg)
        cell.border=bd()
        cell.alignment=right() if c>=2 else left()

# 개선 목표 섹션
R4 = len(bep_rows)+5
merge_wc(ws4,R4,1,R4,5,'▶ 개선 목표 (스마트팜 고도화 방향)',bold=True,size=11,color=WHITE,bg=GREEN_D,border=True)
R4+=1
imp_heads=['개선 항목','현재','목표','개선 효과 (예상)','우선순위']
for i,h in enumerate(imp_heads):
    wc(ws4,R4,i+1,h,bold=True,size=9,color=WHITE,bg=GREEN_D,align=center())
R4+=1
improvements=[
    ('수확량 증대','92,176 kg/53a','105,000 kg','수입 +7,400만원','★★★'),
    ('단가 향상 (출하처 다변화)','5,827원/kg','6,500원/kg','수입 +6,200만원','★★★'),
    ('고용노동 절감 (자동화)','12,295 h','9,000 h','비용 -3,500만원','★★★'),
    ('에너지비 절감','41,410,310원','33,000,000원','비용 -840만원','★★'),
    ('기타재료비 절감','50,121,104원','42,000,000원','비용 -810만원','★★'),
    ('소득률 향상','35.9%','42%↑','소득 +6,600만원','★★★'),
]
for row in improvements:
    for c,v in enumerate(row):
        cell=ws4.cell(row=R4,column=c+1,value=v)
        cell.font=Font(name='맑은 고딕',size=9)
        cell.fill=hd_fill(GREEN_L if c==0 else YELLOW)
        cell.border=bd()
        cell.alignment=left() if c<=1 else right()
    R4+=1

# ── 저장 ────────────────────────────────────────────────────────
out = r'C:\Users\IIamHub3\Documents\네이트온 받은 파일\12.스마트팜_경영성과체계화_방울토마토_장호림.xlsx'
wb.save(out)
print(f"저장 완료: {out}")
