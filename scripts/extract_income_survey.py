"""소득조사표 추출 스크립트 — 농진청 이암허브 스마트팜 패널

입력 : 기초자료/.../환경생육소득(이암허브)/스마트팜 {작목} 소득조사표.zip
       각 ZIP에 농가별 XLSX 5건 포함

출력 : api/data/income_survey.json

구조:
{
  "딸기": {
    "cost_per_m2": {
      "seed":          float,   # 종묘비 (원/m²)
      "fertilizer":    float,   # 비료비 (유기+무기)
      "pesticide":     float,   # 농약비
      "utility":       float,   # 수도광열비 (전기+난방+수도 합산)
      "labor":         float,   # 인건비 (고용+자가)
      "depreciation":  float,   # 상각비
      "other":         float,   # 기타재료비 + 소농구비 + 수리비 + 기타비용
    },
    "n_farms":  int,
    "source":   "xlsx_data" | "rda_reference",
    "note":     str
  },
  ...
}

실행:
    python scripts/extract_income_survey.py
"""
from __future__ import annotations

import io
import json
import zipfile
from pathlib import Path
from typing import Optional

import openpyxl

# ── 경로 설정 ─────────────────────────────────────────────────────────────────
ROOT     = Path(__file__).parent.parent
DATA_DIR = ROOT / "기초자료" / "농진청빅데이터-20260509T093516Z-3-001" / "농진청빅데이터" / "환경생육소득(이암허브)"
OUT_PATH = ROOT / "api" / "data" / "income_survey.json"

# ── 대상 작목 → ZIP 파일명 ─────────────────────────────────────────────────────
CROP_ZIPS: dict[str, str] = {
    "딸기":       "스마트팜 딸기 소득조사표.zip",
    "방울토마토": "스마트팜 방울토마토 소득조사표.zip",
    "완숙토마토": "스마트팜 완숙토마토 소득조사표.zip",
    "참외":       "스마트팜 참외 소득조사표.zip",
}

# ── 농진청 농업경영 통계 참고값 (2018~2022 평균, 원/10a 기준 → 원/m²로 환산)
# 출처: 농촌진흥청 농축산물 생산비 조사 (https://www.rda.go.kr)
# 10a = 1,000m²  →  값÷1000 = 원/m²
RDA_REFERENCE: dict[str, dict[str, float]] = {
    "딸기": {
        "seed":         800_000 / 1000,   # 800원/m²
        "fertilizer":   180_000 / 1000,   # 180원/m²
        "pesticide":    280_000 / 1000,   # 280원/m²
        "utility":      450_000 / 1000,   # 450원/m² (전기+난방 포함)
        "labor":      2_800_000 / 1000,   # 2,800원/m²
        "depreciation": 162_000 / 1000,   # 162원/m²
        "other":        500_000 / 1000,   # 500원/m²
    },
    "방울토마토": {
        "seed":         600_000 / 1000,
        "fertilizer":   220_000 / 1000,
        "pesticide":    260_000 / 1000,
        "utility":      380_000 / 1000,
        "labor":      2_200_000 / 1000,
        "depreciation": 150_000 / 1000,
        "other":        400_000 / 1000,
    },
    "완숙토마토": {
        "seed":         500_000 / 1000,
        "fertilizer":   200_000 / 1000,
        "pesticide":    240_000 / 1000,
        "utility":      350_000 / 1000,
        "labor":      2_100_000 / 1000,
        "depreciation": 145_000 / 1000,
        "other":        400_000 / 1000,
    },
    "참외": {
        "seed":         900_000 / 1000,
        "fertilizer":   190_000 / 1000,
        "pesticide":    300_000 / 1000,
        "utility":      320_000 / 1000,
        "labor":      2_500_000 / 1000,
        "depreciation": 155_000 / 1000,
        "other":        450_000 / 1000,
    },
}

# ── 소득분석2 시트의 행 위치 ──────────────────────────────────────────────────
# (row_index, col_index_total, col_index_per10a)
FIELD_ROWS = {
    "환산면적_a":   (3,  3, None),  # 환산면적(a) — per-10a 환산 기준
    "seed":        (8,  2, 3),     # 종자(묘)비
    "organic_fert":(9,  2, 3),     # 비료 유기질
    "inorganic_fert":(10,2, 3),    # 비료 무기질
    "utility":     (17, 2, 3),     # 수도광열비
    "other_mat":   (18, 2, 3),     # 기타재료비
    "small_equip": (19, 2, 3),     # 소농구비
    "depreciation":(20, 2, 3),     # 상각비(대농구.시설)
    "repair":      (23, 2, 3),     # 수리유지비
    "other_cost":  (24, 2, 3),     # 기타비용
    "contract":    (26, 2, 3),     # 위탁영농비
    "rent":        (27, 2, 3),     # 임차료
    "hired_labor": (28, 2, 3),     # 고용노동비
    "own_labor":   (30, 2, 3),     # 자가노동비
}


def _safe_float(v) -> Optional[float]:
    try:
        f = float(v)
        return f if f > 0 else None
    except (TypeError, ValueError):
        return None


def _decode_zip_name(raw: bytes) -> str:
    try:
        return raw.decode("utf-8")
    except UnicodeDecodeError:
        return raw.decode("euc-kr", errors="replace")


def extract_from_xlsx(raw: bytes) -> Optional[dict]:
    """XLSX 바이트 데이터에서 소득분석2 시트 데이터 추출.

    Returns None if the workbook is a blank template (all cost fields = 0).
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        print(f"    [warn] XLSX 열기 실패: {e}")
        return None

    sheet_name = "소득분석2"
    if sheet_name not in wb.sheetnames:
        print(f"    [warn] '{sheet_name}' 시트 없음. sheets={wb.sheetnames[:5]}")
        return None

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    def cell(r: int, c: int):
        try:
            return rows[r][c]
        except IndexError:
            return None

    # 환산면적 (a 단위: 1a = 100m²)
    area_a = _safe_float(cell(FIELD_ROWS["환산면적_a"][0], FIELD_ROWS["환산면적_a"][1]))
    if not area_a:
        area_a = 10.0  # 기본값: 10a = 1000m²

    area_m2 = area_a * 100.0

    # 필드별 합계(원) 읽기 — col 2(total) 우선, 없으면 per-10a(col 3) × 면적 환산
    def read_field(key: str) -> Optional[float]:
        r, c_tot, c_per = FIELD_ROWS[key]
        v = _safe_float(cell(r, c_tot))
        if v is not None:
            return v
        if c_per is not None:
            v_per = _safe_float(cell(r, c_per))
            if v_per is not None:
                return v_per * area_a  # per-10a × 10a수
        return None

    data: dict = {}
    for k in FIELD_ROWS:
        if k == "환산면적_a":
            continue
        data[k] = read_field(k)

    # 실제 입력된 항목이 하나도 없으면 blank template
    cost_vals = [v for k, v in data.items()
                 if k not in ("depreciation",) and v is not None and v > 0]
    if not cost_vals:
        return None  # 빈 템플릿

    # 원/m² 로 정규화
    def to_per_m2(v: Optional[float]) -> float:
        return round((v or 0) / area_m2, 2)

    return {
        "area_m2":    area_m2,
        "seed":       to_per_m2(data.get("seed")),
        "fertilizer": to_per_m2((data.get("organic_fert") or 0) + (data.get("inorganic_fert") or 0)),
        "utility":    to_per_m2(data.get("utility")),
        "labor":      to_per_m2((data.get("hired_labor") or 0) + (data.get("own_labor") or 0)),
        "depreciation": to_per_m2(data.get("depreciation")),
        "other":      to_per_m2(
            (data.get("other_mat") or 0) +
            (data.get("small_equip") or 0) +
            (data.get("repair") or 0) +
            (data.get("other_cost") or 0)
        ),
    }
    # pesticide is in separate sheet; skip here (will use pesticide sheet if needed)


def extract_pesticide_from_xlsx(raw: bytes, area_m2: float) -> Optional[float]:
    """농약비 시트에서 합계(원/m²) 추출."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception:
        return None
    if "농약비" not in wb.sheetnames:
        return None
    ws = wb["농약비"]
    rows = list(ws.iter_rows(values_only=True))
    # Row 23: 합계 행 (col 9=살포횟수, col 10=비용)
    try:
        total_cost = _safe_float(rows[23][10])
    except IndexError:
        total_cost = None
    if total_cost and area_m2 > 0:
        return round(total_cost / area_m2, 2)
    return None


def process_crop_zip(crop: str, zip_path: Path) -> dict:
    """ZIP 파일에서 농가별 XLSX 읽어 작목 평균 비용 계산."""
    if not zip_path.exists():
        print(f"  [skip] ZIP 없음: {zip_path.name}")
        return {"cost_per_m2": RDA_REFERENCE[crop], "n_farms": 0,
                "source": "rda_reference", "note": "ZIP 파일 없음"}

    farm_records: list[dict] = []

    with zipfile.ZipFile(zip_path) as zf:
        for info in zf.infolist():
            # filename may be garbled (EUC-KR in ZIP decoded as CP437/Latin-1)
            # just check suffix without re-encoding to avoid codec errors
            if not info.filename.lower().endswith(".xlsx"):
                continue
            name = info.filename
            raw = zf.read(info.filename)
            record = extract_from_xlsx(raw)
            if record is not None:
                # pesticide from 농약비 sheet
                pest = extract_pesticide_from_xlsx(raw, record["area_m2"])
                record["pesticide"] = pest or 0.0
                farm_records.append(record)
                print(f"    {Path(name).name}: 면적={record['area_m2']:.0f}m², "
                      f"수도광열={record['utility']:.0f}원/m², 인건비={record['labor']:.0f}원/m²")
            else:
                print(f"    {Path(name).name}: 빈 템플릿 - 스킵")

    if not farm_records:
        print(f"  [{crop}] 실제 데이터 없음 → RDA 통계 참고값 사용")
        return {
            "cost_per_m2": RDA_REFERENCE[crop],
            "n_farms": 0,
            "source": "rda_reference",
            "note": "XLSX 모두 빈 템플릿 — RDA 농업경영 통계(2018-2022) 참고값 적용",
        }

    # 농가 평균
    n = len(farm_records)
    keys = ["seed", "fertilizer", "pesticide", "utility", "labor", "depreciation", "other"]
    avg: dict[str, float] = {}
    for k in keys:
        vals = [r[k] for r in farm_records if r.get(k) is not None]
        avg[k] = round(sum(vals) / len(vals), 2) if vals else RDA_REFERENCE[crop].get(k, 0.0)

    print(f"  [{crop}] 실측 {n}개 농가 평균 계산 완료")
    return {
        "cost_per_m2": avg,
        "n_farms": n,
        "source": "xlsx_data",
        "note": f"소득조사표 {n}개 농가 평균 (빈 템플릿 제외)",
    }


def main():
    print("=" * 60)
    print("소득조사표 추출 시작")
    print("=" * 60)

    result: dict = {}

    for crop, zip_name in CROP_ZIPS.items():
        print(f"\n[{crop}]")
        zip_path = DATA_DIR / zip_name
        result[crop] = process_crop_zip(crop, zip_path)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"\n[완료] 저장: {OUT_PATH}")
    for crop, info in result.items():
        print(f"  [{crop}] source={info['source']}, n={info['n_farms']}")
        cost = info["cost_per_m2"]
        total = sum(cost.values())
        print(f"         총경영비 ~{total:.0f}원/m2 | 인건비={cost.get('labor', 0):.0f}, 광열비={cost.get('utility', 0):.0f}")


if __name__ == "__main__":
    main()
