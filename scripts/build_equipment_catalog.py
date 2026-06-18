# -*- coding: utf-8 -*-
"""
KAASA smartfarmingsight — 시설 기자재 평가 카탈로그 빌더 (v2 분류체계)
======================================================================
'스마트팜기자재정리_v16_..._분류완료.xlsx' (평가완료 DB)를 읽어
공종별 분류체계 + 5등급(A~E) 평가 + 제품 카탈로그 JSON 을 생성한다.

지속 업그레이드: 새 평가본이 나오면 --src 만 바꿔 재실행하면 전량 갱신(멱등).

산출:
  api/data/reference/equipment_taxonomy_v2.json  공종별 분류체계(건축/기자재 3계층)
  api/data/reference/equipment_catalog.json      평가완료 제품 카탈로그(등급 포함)
  api/data/reference/evaluation_criteria.json    12항목 5등급 평가기준
  api/data/reference/_catalog_meta.json          출처·버전·집계
"""
import argparse, json, os, re
from collections import Counter, defaultdict
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(ROOT, "api", "data", "reference")

CRITERIA_KEYS = [
    "자동화 수준", "ICT 연동성", "정밀제어 가능성", "초기 설치 비용", "유지관리 비용",
    "설치의 용이성", "유지보수 편의성", "타 시스템과 호환성",
    "모듈화 및 업그레이드 가능성", "인터페이스(UI)", "교육/지원 서비스",
]


def _n(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in ("None", "nan") else s


def _num(v):
    s = _n(v).replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def build_taxonomy(wb):
    """공종별구분 시트 → 건축/기자재 3계층."""
    ws = wb["공종별구분"]
    rows = [[_n(c) for c in r] for r in ws.iter_rows(values_only=True)]
    tree = defaultdict(lambda: defaultdict(list))   # 구분 → 중분류 → [{품목, 공종}]
    cur_g, cur_m = "", ""
    for r in rows[2:]:
        # 컬럼: [_, 구분, 중분류, 내역(품목), 공종별구분]
        g, m, item, gong = r[1], r[2], r[3], r[4]
        if g:
            cur_g = g
        if m:
            cur_m = m
        if item:
            tree[cur_g][cur_m].append({"item": item, "gong": gong})
    out = []
    for g, mids in tree.items():
        out.append({
            "group": g,
            "mids": [{"mid": m, "items": items} for m, items in mids.items()],
        })
    return out


def build_criteria(wb):
    ws = wb["평가기준"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [_n(c) for c in rows[0]]
    out = []
    for r in rows[1:]:
        r = [_n(c) for c in r]
        if not r[0]:
            continue
        out.append({
            "item": r[0],
            "grades": {f"{i}등급": r[i] for i in range(1, 6) if i < len(r)},
            "source": r[6] if len(r) > 6 else "",
        })
    return out


def build_catalog(wb):
    ws = wb["종합평가"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [_n(c) for c in rows[0]]
    ix = {h: i for i, h in enumerate(hdr)}

    def g(r, name):
        return _n(r[ix[name]]) if name in ix else ""

    out = []
    gubun_c, grade_c, gong_c = Counter(), Counter(), Counter()
    makers = set()
    for r in rows[1:]:
        gubun = g(r, "구분")
        mtype = g(r, "기종명")
        maker = g(r, "업체명")
        form = g(r, "형식명")
        if not (gubun or mtype or maker):
            continue
        scores = {k: _num(r[ix[k]]) for k in CRITERIA_KEYS if k in ix}
        rec = {
            "gubun": gubun,                  # 대분류(구동기/측정기/제어기/농작업기/기타)
            "model_type": mtype,             # 기종명
            "detail": g(r, "기종 세부"),
            "maker": maker,                  # 업체명
            "form_name": form,               # 형식명(모델)
            "price": _num(r[ix["가격"]]) if "가격" in ix else None,  # 만원 추정
            "spec": g(r, "규격"),
            "kc": g(r, "kc 인증 여부"),
            "function": g(r, "주요기능"),
            "grade": g(r, "종합 등급"),       # A~E
            "score100": _num(r[ix["환산 점수(100점)"]]) if "환산 점수(100점)" in ix else None,
            "gong": g(r, "공종별 구분"),       # 공종
            "scores": {k: v for k, v in scores.items() if v is not None},
        }
        out.append(rec)
        if gubun:
            gubun_c[gubun] += 1
        if rec["grade"]:
            grade_c[rec["grade"]] += 1
        if rec["gong"]:
            gong_c[rec["gong"]] += 1
        if maker:
            makers.add(maker)
    stats = {
        "products": len(out),
        "makers": len(makers),
        "by_gubun": dict(gubun_c),
        "by_grade": dict(sorted(grade_c.items())),
        "by_gong": dict(gong_c.most_common()),
    }
    return out, stats


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", required=True)
    a = ap.parse_args()
    wb = openpyxl.load_workbook(a.src, read_only=True, data_only=True)

    taxonomy = build_taxonomy(wb)
    criteria = build_criteria(wb)
    catalog, stats = build_catalog(wb)

    os.makedirs(OUT_DIR, exist_ok=True)

    def dump(fn, obj):
        with open(os.path.join(OUT_DIR, fn), "w", encoding="utf-8") as f:
            json.dump(obj, f, ensure_ascii=False, indent=2)

    src_base = os.path.basename(a.src)
    ver = re.search(r"_v(\d+)", src_base)
    meta = {
        "source": src_base,
        "source_version": ("v" + ver.group(1)) if ver else "",
        "classification": "공종별 분류체계(건축/기자재) + 5등급(A~E) 12항목 평가",
        "stats": stats,
        "criteria_items": [c["item"] for c in criteria],
    }
    dump("equipment_taxonomy_v2.json", taxonomy)
    dump("equipment_catalog.json", catalog)
    dump("evaluation_criteria.json", criteria)
    dump("_catalog_meta.json", meta)
    print("[OK] catalog built ->", OUT_DIR)
    print(json.dumps(meta, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
