# -*- coding: utf-8 -*-
"""20농가 소득조사표 '농가' 시트에서 실제 자산등록부 추출.
발견: 대농기구·영농시설 상세표에 신조가격(취득가)·내용년수·사용년수·규격이 실입력.
→ out/asset_register.json (자산 raw + 농가/작목 집계)."""
import openpyxl, os, json, re
from pathlib import Path

BASE = Path(r"C:/hub/smartfarm-mvp/빅데이터생성형AI수익성향상/extracted_data/농진청데이터")
FILES = {
    "참외": ["참외/스마트팜 참외 소득조사표/스마트팜 경영데이터 조사표_참외 {n}_검증.xlsx",
             ["강석구","유준상","이명화","정준교","조원호"]],
    "방울토마토": ["방울토마토/스마트팜 방울토마토 소득조사표/스마트팜 경영데이터 조사표_방울토마토 {n}_v.3_검증.xlsx",
             ["박경종","박상열","장호림","진화","한태웅"]],
    "딸기": ["딸기/딸기 소득조사표/{f}",
             ["스마트팜 경영데이터 조사표_딸기 류필영_보완_v.3_검증.xlsx",
              "스마트팜 경영데이터 조사표_딸기 문병도_검증.xlsx",
              "스마트팜 경영데이터 조사표_딸기 이병권_검증.xlsx",
              "스마트팜 경영데이터 조사표_딸기 이시성_검증.xlsx",
              "스마트팜 경영데이터 조사표_딸기 류창영_검증.xlsx"]],
    "완숙토마토": ["토마토/스마트팜 완숙토마토 소득조사표/스마트팜 경영데이터 조사표_완숙토마토 {n}_v.3_검증.xlsx",
             ["김선환","모천운","이영재","이태희","최양언"]],
}

def num(v):
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, str):
        s = re.sub(r"[^\d.]", "", v)
        try: return float(s) if s else None
        except: return None
    return None

def extract_section(ws, marker):
    """marker 'o 대농기구' / 'o 영농시설' 행 찾아 자산 파싱."""
    srow = None; deprec_total = None
    for row in ws.iter_rows():
        for c in row:
            if c.value and isinstance(c.value, str) and c.value.strip() == marker:
                srow = c.row
                deprec_total = num(ws.cell(row=c.row, column=13).value)  # M열
                break
        if srow: break
    if not srow: return [], None
    items = []
    r = srow + 4  # o마커 → +2 헤더 → +2 첫 자산
    while r < srow + 70:
        name = ws.cell(row=r, column=2).value
        if isinstance(name, str) and name.strip().startswith("주"): break
        sinjo = num(ws.cell(row=r, column=8).value)     # H 신조가격
        life  = num(ws.cell(row=r, column=12).value)    # L 내용년수
        used  = num(ws.cell(row=r, column=16).value)    # P 사용년수
        spec  = ws.cell(row=r+1, column=3).value        # C of ( ) row = 규격/제조사
        if isinstance(name, str) and name.strip() and sinjo and sinjo > 0:
            items.append({
                "asset": name.strip(),
                "acq_krw": int(sinjo),
                "life_yr": int(life) if life else None,
                "used_yr": int(used) if used else None,
                "spec": str(spec).strip() if spec not in (None, "") else None,
                "annual_deprec": round(sinjo / life) if life else None,
            })
        r += 2
    return items, deprec_total

register = {"_note": "20농가 소득조사표 '농가' 시트 실 자산등록부. acq_krw=신조가격(취득가), life_yr=내용년수, spec=규격/제조사(( )칸). 감가=신조가/내용연수.",
            "crops": {}}
allrows = []
for crop, (tmpl, names) in FILES.items():
    farms = []
    for n in names:
        if "{f}" in tmpl: rel = tmpl.format(f=n); farmname = n.split("_")[2] if len(n.split("_"))>2 else n
        else: rel = tmpl.format(n=n); farmname = n
        f = BASE / rel
        if not f.exists():
            print(f"  ✗ 없음: {f}"); continue
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb["농가"]
        eq, eq_dep = extract_section(ws, "o 대농기구")
        fac, fac_dep = extract_section(ws, "o 영농시설")
        for it in eq: it["category"] = "대농기구"
        for it in fac: it["category"] = "영농시설"
        assets = eq + fac
        cap = sum(a["acq_krw"] for a in assets)
        dep = sum(a["annual_deprec"] for a in assets if a["annual_deprec"])
        farms.append({"farm": farmname, "n_assets": len(assets),
                      "capex_total_krw": cap, "annual_deprec_gross": dep,
                      "reported_deprec": round((eq_dep or 0)+(fac_dep or 0)),
                      "assets": assets})
        for a in assets: allrows.append({"crop": crop, "farm": farmname, **a})
    n = len(farms)
    register["crops"][crop] = {
        "n_farms": n,
        "capex_mean_krw": round(sum(x["capex_total_krw"] for x in farms)/n) if n else 0,
        "capex_min": min((x["capex_total_krw"] for x in farms), default=0),
        "capex_max": max((x["capex_total_krw"] for x in farms), default=0),
        "annual_deprec_mean": round(sum(x["annual_deprec_gross"] for x in farms)/n) if n else 0,
        "farms": farms,
    }

Path("out").mkdir(exist_ok=True)
Path("out/asset_register.json").write_text(json.dumps(register, ensure_ascii=False, indent=1), encoding="utf-8")

# 요약 출력
print("작목 | 농가 | 평균CAPEX(취득가합) | 평균 연감가(gross) | CAPEX범위")
for crop, d in register["crops"].items():
    print(f"  {crop:6} | {d['n_farms']} | {d['capex_mean_krw']:>14,} | {d['annual_deprec_mean']:>11,} | {d['capex_min']:,}~{d['capex_max']:,}")
# 규격/제조사 채움률
spec_n = sum(1 for a in allrows if a.get("spec"))
print(f"\n총 자산행 {len(allrows)} · 규격/제조사 기입 {spec_n} ({100*spec_n//max(1,len(allrows))}%)")
print("규격 샘플:")
for a in [x for x in allrows if x.get("spec")][:12]:
    print(f"  [{a['crop']}/{a['farm']}] {a['asset']} · 취득 {a['acq_krw']:,} · {a['life_yr']}년 · 규격 {a['spec']}")
