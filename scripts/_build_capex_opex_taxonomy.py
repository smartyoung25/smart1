# -*- coding: utf-8 -*-
"""딸기 스마트팜 CAPEX/OPEX 체계화·계층화 (.xlsx)

근거: 농진청 소득조사표(딸기 류창영·900㎡) — 상각비 대농구 675,000 + 영농시설 783,000,
      환경관리시설 제어현황(복합환경제어·관수관비·보온/차광스크린·환기·CO2·난방),
      OPEX 시트(수리유지비=대농기구/영농시설 분리, 수도광열=전기/경유/물, 임차료=토지/대농기구/영농시설).

산출: CAPEX 3계층(대분류>중분류>세부품목) + 업체·성능·취득가·내용연수·잔존율·정액감가상각(수식)
      + 연동 OPEX. 조사표는 집계 상각비만 있어 per-asset 취득가는 「입력」(견적서 연동)·내용연수는
      법인세법 시행규칙 별표·농진청 농기계 내용연수 표준을 적용(비고에 명시).

사용: PYTHONPATH=C:/smart_farm python scripts/_build_capex_opex_taxonomy.py
"""
from __future__ import annotations
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "out" / "스마트팜_CAPEX_OPEX_체계화_4작목.xlsx"
FONT = "맑은 고딕"

# ── 20농가 실 자산등록부 (scripts/_extract_asset_register.py 산출) ────────────
# '농가' 시트 대농기구·영농시설 상세표에서 신조가격(취득가)·내용년수·사용년수·규격(사양)을 실추출.
# ★ 앞선 "취득가 미수집" 판단 정정 — 조사표 '농가' 시트에 자산별 실 취득가·사양이 존재한다.
_REG_PATH = ROOT / "out" / "asset_register.json"
REG = json.loads(_REG_PATH.read_text(encoding="utf-8")) if _REG_PATH.exists() else {"crops": {}}

# ── 작목별 실 시설 구성 (농진청 소득조사표 '농가' 시트 제어현황 설치유무) ────────
# ○ 설치 · × 미설치 · '고장' 원문. 소득분석2 의 집계 상각비(675k/783k·홍길동)는
# 샘플 템플릿이라 작목 공통이며 실 취득가가 아니다 — 실 차이는 아래 시설 구성이다.
CROP_FACILITY = {
    "항목": ["일중천장", "이중천장", "측창(환기)", "천정 보온스크린", "측면 보온스크린", "차광스크린", "관수·관비장치"],
    "딸기 (류창영)":     ["○", "×", "○", "○", "○", "○", "○"],
    "방울토마토 (박경종)": ["○", "×", "○", "○", "×", "○", "○"],
    "완숙토마토 (김선환)": ["○", "×", "×", "○", "×", "○", "×"],
    "참외 (강석구)":     ["○", "×", "○", "×", "×", "×", "고장"],
}
CROP_META = {  # (재배유형, 난방에너지원, 생산량kg)
    "딸기 (류창영)":     ("촉성(1)", "전기+기타(2,10)", "23,516"),
    "방울토마토 (박경종)": ("촉성(1)", "기타(10)", "65,464"),
    "완숙토마토 (김선환)": ("촉성(1)", "등유+제습기(2,10)", "198,797"),
    "참외 (강석구)":     ("반촉성(2)", "심야전기(9)", "50,000"),
}

# 색
C_TITLE = "1C5A3A"; C_HDR = "2F9A62"; C_L1 = "DCE6F1"; C_SOFT = "F4F7F4"; C_INK = "16211B"
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def _f(sz=10, b=False, color=C_INK): return Font(name=FONT, size=sz, bold=b, color=color)
def _fill(c): return PatternFill("solid", fgColor=c)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEF = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIG = Alignment(horizontal="right", vertical="center")

# ── CAPEX 계층 데이터 ────────────────────────────────────────────────────────
# (중분류, 세부품목, 업체예시, 성능·사양예시, 취득가액, 내용연수, 잔존율, 연동OPEX, 비고)
# 취득가액 0 = 「견적서 입력」. 내용연수=표준(법인세법 별표·농진청 농기계 기준).
CAPEX = {
    "① 영농시설 (구조·피복·보온·차광)": [
        ("온실 구조체", "철골 골조·기초", "(예) 그린플러스·명성GT", "단동/연동 비닐온실, 내재해형", 0, 12, 10, "수리유지비(영농시설)·임차료(영농시설)", "비닐온실 철골 10~15년"),
        ("온실 구조체", "유리/경질판 온실", "(예) 대한하우스", "벤로형 유리온실", 0, 20, 10, "수리유지비(영농시설)", "유리온실 15~20년"),
        ("피복", "피복필름(장기성 PO)", "(예) 일신화학", "PO 0.15mm, 광투과 90%+", 0, 3, 0, "당해 자재비(소모성)", "★소모성 1~3년 — 자산화 여부 확인"),
        ("보온", "천정 보온스크린", "(예) 룩스퍼스", "다겹 보온커튼, 보온율 60%+", 0, 6, 10, "수리유지비(영농시설)·전기(개폐)", "스크린 5~7년"),
        ("보온", "측면 보온스크린", "(예) 룩스퍼스", "다겹, 자동개폐", 0, 6, 10, "수리유지비(영농시설)·전기", "스크린 5~7년"),
        ("차광", "차광 스크린", "(예) 스벤손", "알루미늄 차광 50~70%", 0, 6, 10, "수리유지비(영농시설)", "스크린 5~7년"),
        ("재배베드", "재배베드(거터)·배지받침", "(예) 그로단·리코", "고설 벤치·NFT 거터", 0, 8, 10, "수리유지비(영농시설)·배지 교체비", "설비 8~10년, 배지는 소모성"),
    ],
    "② 환경관리 기자재 (제어·설비)": [
        ("복합환경제어", "복합환경제어 시스템", "(예) 프리바·나래트렌드", "제어반+SW, 온·습·CO2·관수 통합", 0, 8, 10, "전기·SW 구독료·수리유지비", "제어설비 7~10년"),
        ("관수·관비", "양액기(관비장치)·펌프", "(예) 네타핌·프리바", "EC/pH 자동제어, 다구역", 0, 8, 10, "전기·물·양액소재비·수리유지비", "관수설비 7~10년"),
        ("난방", "난방기(온풍/온수)", "(예) 대성하이텍", "경유/가스/히트펌프", 0, 9, 10, "수도광열비(경유·가스·전기)·수리유지비", "난방기 8~10년"),
        ("CO2 시비", "CO2 발생·공급 설비", "(예) 코비시스템", "액화CO2/연소식 시비", 0, 8, 10, "CO2 원료비·전기", "8~10년"),
        ("환기", "측창/천창 개폐 모터", "(예) 리코엔지니어링", "감속모터·랙피니언", 0, 8, 10, "전기·수리유지비", "8~10년"),
        ("계측·센서", "환경센서(온습도·일사·CO2·EC/pH)", "(예) 센서웨이", "무선/유선 다지점", 0, 5, 10, "전기·교정비·소모품(전극)", "센서 5년"),
    ],
    "③ 대농기구 (운반·방제·관리)": [
        ("운반", "전동 운반차·모노레일", "(예) 아그로텍", "적재 300kg급", 0, 7, 10, "수리유지비(대농기구)·전기/유류·임차료(대안)", "운반기구 5~8년"),
        ("방제", "방제기(동력분무·연무기)", "(예) 아세아텍", "동력 SS/연무", 0, 6, 10, "수리유지비(대농기구)·유류·약제비", "방제기 5~7년"),
        ("관리", "예초기·기타 관리기구", "(예) 계양", "휴대형", 0, 5, 10, "수리유지비(대농기구)·유류", "5~7년"),
    ],
    "④ 소모성 자재 (당해비용 · 감가상각 대상 아님)": [
        ("소모성", "배지(코코피트·암면)", "(예) 그로단", "슬라브/포트", 0, 1, 0, "→ OPEX(기타재료비)", "★자산 아님 — 당해 비용처리"),
        ("소모성", "양액 원소재·소독제", "-", "-", 0, 1, 0, "→ OPEX(기타재료비)", "★당해 비용처리"),
        ("소모성", "소농구(호미·낫·전정가위 등)", "-", "-", 0, 1, 0, "→ OPEX(소농구비)", "★조사표 소농구비 항목"),
    ],
}

wb = Workbook()

# ══ Sheet1: CAPEX 자산 계층 등록부 ══════════════════════════════════════════
ws = wb.active; ws.title = "CAPEX 계층 등록부"
COLS = ["대분류", "중분류", "세부품목", "업체(제조사)", "성능·사양", "취득가액(원)",
        "내용연수(년)", "잔존율(%)", "연 감가상각(정액,원)", "연동 OPEX", "비고"]
W = [17, 13, 20, 18, 22, 13, 9, 8, 15, 24, 20]
for i, w in enumerate(W): ws.column_dimensions[get_column_letter(i+1)].width = w

ws.merge_cells("A1:K1")
ws["A1"] = "스마트팜 CAPEX 자산 계층 등록부 (3계층 · 정액 감가상각 · 시설작목 공통 프레임)"
ws["A1"].font = _f(14, True, "FFFFFF"); ws["A1"].fill = _fill(C_TITLE); ws["A1"].alignment = CEN
ws.row_dimensions[1].height = 26
ws.merge_cells("A2:K2")
ws["A2"] = ("근거: 농진청 스마트팜 소득조사표(딸기·방울·완숙토마토·참외 4작목·20농가). "
            "★ '소득분석2'의 집계 상각비(675k/783k·'홍길동')는 샘플 템플릿이나, 조사표 '농가' 시트에는 자산별 "
            "신조가격(취득가)·내용년수·규격(사양)이 실입력돼 있다 → '실 자산등록부(20농가)'·'작목별 CAPEX 실측' 시트 참조. "
            "본 시트는 표준 계층 프레임(내용연수=법인세법 별표·농진청 표준, 잔존율 10%)이며 연 감가상각 = 취득가액×(1−잔존율)÷내용연수. "
            "업체(제조사)는 조사표에 거의 미기재(사양은 45% 기입) — 견적서 연동으로 보강 필요.")
ws["A2"].font = _f(8, False, "595959"); ws["A2"].alignment = LEF
ws.row_dimensions[2].height = 30

hr = 3
for i, c in enumerate(COLS):
    cell = ws.cell(hr, i+1, c); cell.font = _f(9, True, "FFFFFF"); cell.fill = _fill(C_HDR)
    cell.alignment = CEN; cell.border = BORDER

r = hr + 1
for l1, items in CAPEX.items():
    l1_start = r
    for it in items:
        mid, item, maker, spec, cost, life, resid, opex, note = it
        vals = [l1, mid, item, maker, spec, cost, life, resid, None, opex, note]
        for ci, v in enumerate(vals):
            cell = ws.cell(r, ci+1, v); cell.border = BORDER; cell.font = _f(9)
            if ci in (5, 8): cell.alignment = RIG; cell.number_format = '#,##0'
            elif ci in (6, 7): cell.alignment = CEN
            else: cell.alignment = LEF
        # 연 감가상각 수식 (소모성=0년이면 상각 없음)
        fcell = ws.cell(r, 9)
        if life and int(life) >= 2:
            fcell.value = f"=F{r}*(1-H{r}/100)/G{r}"
        else:
            fcell.value = 0
        fcell.number_format = '#,##0'
        r += 1
    # 대분류 셀 병합
    ws.merge_cells(start_row=l1_start, start_column=1, end_row=r-1, end_column=1)
    c = ws.cell(l1_start, 1); c.fill = _fill(C_L1); c.font = _f(10, True, C_TITLE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 합계행
ws.cell(r, 3, "① ~ ③ 감가상각 합계 (소모성 제외)").font = _f(9, True)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
tot = ws.cell(r, 9)
# 소모성(④) 행 제외: ④는 마지막 3행 → 상각합계는 그 위까지
last_asset = r - 1 - len(CAPEX["④ 소모성 자재 (당해비용 · 감가상각 대상 아님)"])
tot.value = f"=SUM(I{hr+1}:I{last_asset})"; tot.number_format = '#,##0'; tot.font = _f(10, True, C_TITLE)
tot.fill = _fill(C_L1); tot.alignment = RIG; tot.border = BORDER
for cc in range(1, 12): ws.cell(r, cc).border = BORDER; ws.cell(r, cc).fill = _fill(C_L1)
sumrow = r

# 조사표 대조행
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
ws.cell(r, 1, "▶ 조사표 템플릿 상각비(대농구 675,000 + 영농시설 783,000 = 1,458,000·샘플값) — 실 취득가 입력 시 위 합계로 대체·대조").font = _f(9, False, "A6431E")
ws.cell(r, 9, 1458000).number_format = '#,##0'; ws.cell(r, 9).font = _f(9, True, "A6431E"); ws.cell(r, 9).alignment = RIG
ws.freeze_panes = "A4"

# ══ Sheet2: OPEX ↔ CAPEX 연동 매트릭스 (3대분류 통합) ══════════════════════
# ★ 과세분 정리: 조사표 13개 OPEX 세목을 3대분류(재료비·경비·노무비)로 통합.
#   수도광열(전기/경유/물)·수리유지(시설/대농기구)는 1행으로 접고 세부 연동은 비고에 명시.
ws2 = wb.create_sheet("OPEX-CAPEX 연동")
for i, w in enumerate([12, 24, 13, 26, 34]): ws2.column_dimensions[get_column_letter(i+1)].width = w
ws2.merge_cells("A1:E1")
ws2["A1"] = "OPEX ↔ CAPEX 연동 매트릭스 (3대분류 통합 · 투자비가 유발하는 운영비)"
ws2["A1"].font = _f(13, True, "FFFFFF"); ws2["A1"].fill = _fill(C_TITLE); ws2["A1"].alignment = CEN
ws2.row_dimensions[1].height = 24
ws2.merge_cells("A2:E2")
ws2["A2"] = ("★ 조사표 13개 OPEX 세목(종자·종묘/무기·유기질비료/농약/수도광열/기타재료/소농구/수리유지/기타/임차/위탁/고용/자가노동)을 "
            "재료비·경비·노무비 3대분류로 통합. 수도광열(전기·연료·물)·수리유지(시설·대농기구)는 1행으로 접고 세부 연동은 비고에 표기.")
ws2["A2"].font = _f(8, False, "595959"); ws2["A2"].alignment = LEF; ws2.row_dimensions[2].height = 26
# (대분류, OPEX 항목(조사표 세목 통합), 성격, 연동 CAPEX, 비고)
OP_HDR = ["OPEX 대분류", "OPEX 항목 (조사표 세목 통합)", "성격", "연동 CAPEX", "산정·비고"]
OP = [
    ("재료비", "종자·종묘 / 비료(무기+유기) / 농약 / 기타재료(배지·양액·피복)", "소모성", "④ 소모성 자재", "무기+유기 비료 통합·당해 비용처리(감가 아님)"),
    ("경비",   "수도광열비 (전기·연료·물)", "에너지", "② 제어·관수·난방·환기·센서", "전기→제어/관수/환기·센서, 연료→난방, 물→관수 (세부 접음)"),
    ("경비",   "수리유지비 (영농시설·대농기구)", "유지보수", "①③ 영농시설·대농기구", "취득가 대비 연 1~3%, 노후시 증가 (시설·기구 통합)"),
    ("경비",   "임차료 (토지·대농기구·영농시설)", "임차(소유 대안)", "①③ 자가소유 대체", "자가=감가상각 / 임차=임차료 — 택1"),
    ("경비",   "소농구비", "소모성 기구", "④ 소농구", "20만원 미만·당해 비용"),
    ("노무비", "고용노임 · 자가노동 · 위탁영농비", "노무", "(자산 무관 · 상충)", "CAPEX 자동화↑ → 노무비↓ 상충관계"),
]
for ci, v in enumerate(OP_HDR):
    cell = ws2.cell(3, ci+1, v); cell.border = BORDER; cell.font = _f(9, True, "FFFFFF"); cell.fill = _fill(C_HDR); cell.alignment = CEN
# 대분류 병합 렌더
r2 = 4; grp_start = 4; prev = OP[0][0]
for idx, (grp, item, kind, capex, note) in enumerate(OP):
    for ci, v in enumerate([grp, item, kind, capex, note]):
        cell = ws2.cell(r2, ci+1, v); cell.border = BORDER; cell.font = _f(9); cell.alignment = LEF
        if idx % 2 == 1: cell.fill = _fill(C_SOFT)
    r2 += 1
# 대분류 셀 병합(재료비/경비/노무비)
def _merge_grp(col, val, r0, r1):
    ws2.merge_cells(start_row=r0, start_column=col, end_row=r1, end_column=col)
    c = ws2.cell(r0, col); c.font = _f(10, True, C_TITLE); c.alignment = CEN; c.fill = _fill(C_L1)
_merge_grp(1, "재료비", 4, 4)
_merge_grp(1, "경비", 5, 8)
_merge_grp(1, "노무비", 9, 9)
ws2.merge_cells(start_row=r2+1, start_column=1, end_row=r2+1, end_column=5)
ws2.cell(r2+1, 1, "※ 핵심: 스마트팜 CAPEX(제어·난방·관수 자동화)는 에너지·유지보수 OPEX를 유발하되 노무비는 절감한다. "
        "투자 의사결정 = 감가상각 + 연동 OPEX 증분 vs 노무·수율 편익. 세목은 조사표 원본을 3대분류로 집계했다.").font = _f(9, False, "595959")

# ══ Sheet3: 내용연수 표준 & 계층 요약 ═══════════════════════════════════════
ws3 = wb.create_sheet("내용연수 표준·요약")
for i, w in enumerate([26, 14, 40]): ws3.column_dimensions[get_column_letter(i+1)].width = w
ws3.merge_cells("A1:C1")
ws3["A1"] = "자산군별 표준 내용연수(수명연한) & 계층 요약"
ws3["A1"].font = _f(13, True, "FFFFFF"); ws3["A1"].fill = _fill(C_TITLE); ws3["A1"].alignment = CEN
LIFE = [
    ["자산군", "내용연수(년)", "근거·비고"],
    ["유리/경질판 온실", "15~20", "법인세법 시행규칙 별표(건물·구축물)"],
    ["비닐온실(철골)", "10~15", "내재해형 기준, 골조 위주"],
    ["피복필름(PO·비닐)", "1~3", "★소모성 — 자산화 여부 회계기준 확인"],
    ["보온·차광 스크린", "5~7", "개폐 반복 마모"],
    ["복합환경제어 시스템", "7~10", "전자설비, SW 업데이트 별도"],
    ["관수·관비장치(양액기)", "7~10", "펌프·밸브 소모부품 별도"],
    ["난방기", "8~10", "농진청 농업기계 내용연수"],
    ["CO2 발생·환기 설비", "8~10", "기계·전자 혼합"],
    ["환경 센서류", "5", "교정·전극 소모"],
    ["대농기구(운반·방제)", "5~8", "농진청 농업기계 기준"],
]
for ri, row in enumerate(LIFE):
    for ci, v in enumerate(row):
        cell = ws3.cell(ri+2, ci+1, v); cell.border = BORDER; cell.font = _f(9, ri == 0, "FFFFFF" if ri == 0 else C_INK)
        cell.alignment = CEN if ci == 1 else LEF
        if ri == 0: cell.fill = _fill(C_HDR); cell.alignment = CEN
        elif ri % 2 == 0: cell.fill = _fill(C_SOFT)
base = len(LIFE) + 4
ws3.merge_cells(start_row=base, start_column=1, end_row=base, end_column=3)
ws3.cell(base, 1, "계층 구조 요약").font = _f(11, True, C_TITLE)
SUMM = [
    "대분류 3(+소모성) : ① 영농시설  ② 환경관리 기자재  ③ 대농기구  (④ 소모성=당해비용)",
    "중분류 : 구조·피복·보온·차광·베드 / 제어·관수·난방·CO2·환기·센서 / 운반·방제·관리",
    "속성 : 업체(제조사) · 성능·사양 · 취득가액 · 내용연수 · 잔존율 · 연 감가상각(정액) · 연동 OPEX",
    "정액 감가상각 = 취득가액 × (1 − 잔존율) ÷ 내용연수  (잔존율 표준 10%)",
    "조사표 대조 : 대농구 675,000 + 영농시설 783,000 = 1,458,000원 (900㎡, 10a당 16,200원)",
    "OPEX 연동 : 시설→수리유지·임차, 제어/관수/난방→수도광열, 소모성→기타재료·소농구",
]
for i, s in enumerate(SUMM):
    ws3.cell(base+1+i, 1, "• " + s).font = _f(9)
    ws3.merge_cells(start_row=base+1+i, start_column=1, end_row=base+1+i, end_column=3)

# ══ Sheet4: 작목별 시설 구성 비교 (실 데이터) ═══════════════════════════════
ws4 = wb.create_sheet("작목별 시설 구성")
crops = list(CROP_FACILITY.keys())[1:]
ncol = 1 + len(crops)
for i in range(ncol): ws4.column_dimensions[get_column_letter(i+1)].width = 18 if i == 0 else 16
ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
ws4["A1"] = "작목별 스마트팜 시설 구성 (조사표 '농가' 시트 제어현황 · 실 데이터)"
ws4["A1"].font = _f(13, True, "FFFFFF"); ws4["A1"].fill = _fill(C_TITLE); ws4["A1"].alignment = CEN
ws4.row_dimensions[1].height = 24
ws4.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
ws4["A2"] = ("★ CAPEX 실 차이 — ○설치 자산만 취득가·감가상각 대상. 아래 ○/✕는 각 작목 표본농가 1곳. "
             "작목별 5농가(총 20농가) 실측 시설완비도 계수: 완숙 0.940·딸기 0.934·방울 0.910·참외 0.712 — 참외만 확연히 낮다.")
ws4["A2"].font = _f(8, False, "595959"); ws4["A2"].alignment = LEF; ws4.row_dimensions[2].height = 24
# 헤더
hr4 = 3
ws4.cell(hr4, 1, "환경관리 시설 항목").font = _f(9, True, "FFFFFF"); ws4.cell(hr4, 1).fill = _fill(C_HDR); ws4.cell(hr4, 1).alignment = CEN; ws4.cell(hr4, 1).border = BORDER
for j, c in enumerate(crops):
    cell = ws4.cell(hr4, j+2, c); cell.font = _f(9, True, "FFFFFF"); cell.fill = _fill(C_HDR); cell.alignment = CEN; cell.border = BORDER
# 항목 행
for i, name in enumerate(CROP_FACILITY["항목"]):
    rr = hr4 + 1 + i
    ic = ws4.cell(rr, 1, name); ic.font = _f(9, True); ic.alignment = LEF; ic.border = BORDER; ic.fill = _fill(C_SOFT)
    for j, c in enumerate(crops):
        v = CROP_FACILITY[c][i]
        cell = ws4.cell(rr, j+2, v); cell.border = BORDER; cell.alignment = CEN
        cell.font = _f(11, True, "2F9A62" if v == "○" else ("A6431E" if v == "고장" else "B0B0B0"))
# 메타 행
base4 = hr4 + 1 + len(CROP_FACILITY["항목"])
for k, lbl in enumerate(["재배유형", "난방에너지원", "생산량(kg)"]):
    rr = base4 + k
    ws4.cell(rr, 1, lbl).font = _f(9, True, C_TITLE); ws4.cell(rr, 1).alignment = LEF; ws4.cell(rr, 1).border = BORDER; ws4.cell(rr, 1).fill = _fill(C_L1)
    for j, c in enumerate(crops):
        cell = ws4.cell(rr, j+2, CROP_META[c][k]); cell.font = _f(9); cell.alignment = CEN; cell.border = BORDER
# 해설
ws4.merge_cells(start_row=base4+4, start_column=1, end_row=base4+4, end_column=ncol)
ws4.cell(base4+4, 1, "※ 설치된 시설(○)만 CAPEX 등록부에 취득가·내용연수를 입력해 감가상각을 계산한다. "
         "'고장'은 재투자(교체 CAPEX) 또는 수리(OPEX) 의사결정 대상. 작목별 감가상각 총액은 시설 구성 차이만큼 달라진다.").font = _f(9, False, "595959")

# ══ Sheet5: 실 자산등록부 (20농가 · 조사표 '농가' 시트 실추출) ═══════════════
ws5 = wb.create_sheet("실 자산등록부(20농가)")
COLS5 = ["작목", "농가", "분류", "자산", "취득가(신조가,원)", "내용연수(년)", "사용년수(년)", "규격·사양(원문)", "연 감가상각(원)"]
W5 = [11, 9, 10, 20, 15, 10, 10, 26, 14]
for i, w in enumerate(W5): ws5.column_dimensions[get_column_letter(i+1)].width = w
ws5.merge_cells("A1:I1")
ws5["A1"] = "실 자산등록부 — 20농가 소득조사표 '농가' 시트 대농기구·영농시설 상세표 실추출"
ws5["A1"].font = _f(13, True, "FFFFFF"); ws5["A1"].fill = _fill(C_TITLE); ws5["A1"].alignment = CEN
ws5.row_dimensions[1].height = 24
ws5.merge_cells("A2:I2")
ws5["A2"] = ("★ 취득가(신조가격)·내용연수·규격(사양)은 조사표에 실입력 — '취득가 미수집' 이전 판단을 정정한다. "
            "연 감가상각 = 취득가 ÷ 내용연수(gross, 조사작목 부담비율 미적용). "
            "업체(제조사)는 조사표 ( )칸에 거의 미기재(사양 위주 45% 기입) — 규격·사양 원문 그대로 표기했다.")
ws5["A2"].font = _f(8, False, "595959"); ws5["A2"].alignment = LEF; ws5.row_dimensions[2].height = 26
for i, c in enumerate(COLS5):
    cell = ws5.cell(3, i+1, c); cell.font = _f(9, True, "FFFFFF"); cell.fill = _fill(C_HDR); cell.alignment = CEN; cell.border = BORDER
r5 = 4
CROP_ORDER = ["딸기", "방울토마토", "완숙토마토", "참외"]
for crop in CROP_ORDER:
    d = REG["crops"].get(crop, {})
    crop_start = r5
    for fm in d.get("farms", []):
        for a in fm["assets"]:
            vals = [crop, fm["farm"], a["category"], a["asset"], a["acq_krw"], a.get("life_yr"),
                    a.get("used_yr"), a.get("spec") or "", a.get("annual_deprec")]
            for ci, v in enumerate(vals):
                cell = ws5.cell(r5, ci+1, v); cell.border = BORDER; cell.font = _f(8.5)
                if ci in (4, 8): cell.alignment = RIG; cell.number_format = '#,##0'
                elif ci in (5, 6): cell.alignment = CEN
                else: cell.alignment = LEF
                if a["category"] == "영농시설": cell.fill = _fill("F2F7F4")
            r5 += 1
    if r5 > crop_start:
        ws5.merge_cells(start_row=crop_start, start_column=1, end_row=r5-1, end_column=1)
        cc = ws5.cell(crop_start, 1); cc.font = _f(10, True, C_TITLE); cc.alignment = CEN; cc.fill = _fill(C_L1)
ws5.freeze_panes = "A4"

# ══ Sheet6: 작목별 CAPEX 실측 집계 ═════════════════════════════════════════
ws6 = wb.create_sheet("작목별 CAPEX 실측")
COLS6 = ["작목", "농가수", "평균 총취득가(원)", "취득가 범위(원)", "평균 연감가(gross,원)", "대표 영농시설 자산(평균 취득가)"]
W6 = [12, 8, 18, 26, 18, 44]
for i, w in enumerate(W6): ws6.column_dimensions[get_column_letter(i+1)].width = w
ws6.merge_cells("A1:F1")
ws6["A1"] = "작목별 CAPEX 실측 집계 (20농가 자산등록부 기반)"
ws6["A1"].font = _f(13, True, "FFFFFF"); ws6["A1"].fill = _fill(C_TITLE); ws6["A1"].alignment = CEN
ws6.row_dimensions[1].height = 24
ws6.merge_cells("A2:F2")
ws6["A2"] = ("★ 조사표 템플릿 상각비(1,458,000·4작목 동일)를 실 취득가로 대체 — 작목별 CAPEX가 실제로 크게 다르다. "
            "취득가 = 농가별 자산 신조가격 합, 연감가 = Σ(신조가÷내용연수). 규모(온실 형태·면적)가 작목 차이의 주동인.")
ws6["A2"].font = _f(8, False, "595959"); ws6["A2"].alignment = LEF; ws6.row_dimensions[2].height = 24
for i, c in enumerate(COLS6):
    cell = ws6.cell(3, i+1, c); cell.font = _f(9, True, "FFFFFF"); cell.fill = _fill(C_HDR); cell.alignment = CEN; cell.border = BORDER
r6 = 4
from collections import defaultdict
for crop in CROP_ORDER:
    d = REG["crops"].get(crop, {})
    agg = defaultdict(list)
    for fm in d.get("farms", []):
        for a in fm["assets"]:
            if a["category"] == "영농시설": agg[a["asset"]].append(a["acq_krw"])
    top = sorted(agg.items(), key=lambda kv: -sum(kv[1]))[:3]
    rep = " · ".join(f"{k}({round(sum(v)/len(v)):,})" for k, v in top)
    vals = [crop, d.get("n_farms", 0), d.get("capex_mean_krw", 0),
            f"{d.get('capex_min',0):,} ~ {d.get('capex_max',0):,}", d.get("annual_deprec_mean", 0), rep]
    for ci, v in enumerate(vals):
        cell = ws6.cell(r6, ci+1, v); cell.border = BORDER; cell.font = _f(9)
        if ci in (2, 4): cell.alignment = RIG; cell.number_format = '#,##0'
        elif ci == 1: cell.alignment = CEN
        else: cell.alignment = LEF
    r6 += 1
ws6.merge_cells(start_row=r6+1, start_column=1, end_row=r6+1, end_column=6)
ws6.cell(r6+1, 1, "※ 완숙토마토 평균 취득가 14.2억(유리온실 평균 17.9억 포함)~참외 2.0억 — 온실 형태·규모 차이가 CAPEX를 좌우한다. "
         "감가상각은 대형 유리온실(내용연수 30년)일수록 취득가 대비 완만하다.").font = _f(9, False, "595959")

BLUE = "0000FF"  # 입력값(산업표준 색) — 바꾸면 수식 자동 재계산

# ══ Sheet7: 시설 완비도 계수 (산식·live) ═══════════════════════════════════
ws7 = wb.create_sheet("시설완비도 계수(산식)")
ws7.sheet_view.showGridLines = False
for col, w in zip("ABCDEF", [22, 11, 12, 12, 12, 12]): ws7.column_dimensions[col].width = w
ws7.merge_cells("A1:F1")
ws7["A1"] = "시설 완비도 계수 — 산식과 계산 (살아있는 수식)"
ws7["A1"].font = _f(13, True, "FFFFFF"); ws7["A1"].fill = _fill(C_TITLE); ws7["A1"].alignment = CEN; ws7.row_dimensions[1].height = 26
ws7.merge_cells("A2:F2")
ws7["A2"] = ("계수 = 0.60(온실구조 공통) + Σ(시설별 가중치 × 설치배수)  ·  설치배수 ○=1.0·고장=0.5·×=0  ·  가중치 합 0.40 → 최대 1.00")
ws7["A2"].font = _f(9, False, "595959"); ws7["A2"].alignment = LEF; ws7.row_dimensions[2].height = 24
FAC_HDR = ["시설 항목", "가중치", "딸기(류창영)", "방울(박경종)", "완숙(김선환)", "참외(강석구)"]
for j, h in enumerate(FAC_HDR):
    c = ws7.cell(4, j+1, h); c.font = _f(9, True, "FFFFFF"); c.fill = _fill(C_HDR); c.alignment = CEN; c.border = BORDER
FAC_ROWS = [
    ("온실구조체(공통)", 0.60, 1, 1, 1, 1), ("일중천장", 0.03, 1, 1, 1, 1), ("이중천장", 0.03, 0, 0, 0, 0),
    ("측창(환기)", 0.07, 1, 1, 0, 1), ("천정 보온스크린", 0.07, 1, 1, 1, 0), ("측면 보온스크린", 0.06, 1, 0, 0, 0),
    ("차광 스크린", 0.06, 1, 1, 1, 0), ("관수·관비장치", 0.08, 1, 1, 0, 0.5),
]
fr0 = 5
for i, (name, w, *muls) in enumerate(FAC_ROWS):
    r = fr0 + i
    ws7.cell(r, 1, name).font = _f(9, i == 0); ws7.cell(r, 1).alignment = LEF; ws7.cell(r, 1).border = BORDER
    ws7.cell(r, 2, w).font = _f(9, False, BLUE); ws7.cell(r, 2).alignment = CEN; ws7.cell(r, 2).border = BORDER; ws7.cell(r, 2).number_format = '0.00'
    for j, v in enumerate(muls):
        cell = ws7.cell(r, 3+j, v); cell.font = _f(9, False, BLUE); cell.alignment = CEN; cell.border = BORDER; cell.number_format = '0.0'
        if i == 0: cell.fill = _fill(C_SOFT)
fr_last = fr0 + len(FAC_ROWS) - 1
rsum = fr_last + 1
ws7.cell(rsum, 1, "가중치 합").font = _f(9, True, C_TITLE); ws7.cell(rsum, 1).alignment = LEF; ws7.cell(rsum, 1).border = BORDER; ws7.cell(rsum, 1).fill = _fill(C_SOFT)
ws7.cell(rsum, 2, f"=SUM(B{fr0}:B{fr_last})").font = _f(9, True); ws7.cell(rsum, 2).alignment = CEN; ws7.cell(rsum, 2).border = BORDER; ws7.cell(rsum, 2).number_format = '0.00'
rfac = rsum + 1
ws7.cell(rfac, 1, "시설 완비도 계수").font = _f(10, True, C_TITLE); ws7.cell(rfac, 1).alignment = LEF; ws7.cell(rfac, 1).border = BORDER; ws7.cell(rfac, 1).fill = _fill("E3EFE7")
ws7.cell(rfac, 2, "").border = BORDER; ws7.cell(rfac, 2).fill = _fill("E3EFE7")
for j, col in enumerate("CDEF"):
    cell = ws7.cell(rfac, 3+j, f"=SUMPRODUCT($B${fr0}:$B${fr_last},{col}{fr0}:{col}{fr_last})")
    cell.font = _f(11, True, C_HDR); cell.alignment = CEN; cell.border = BORDER; cell.fill = _fill("E3EFE7"); cell.number_format = '0.000'
ws7.merge_cells(start_row=rfac+2, start_column=1, end_row=rfac+2, end_column=6)
ws7.cell(rfac+2, 1, "※ 위 계수는 각 작목 표본농가 1곳에 산식을 적용(딸기0.97·방울0.91·완숙0.76·참외0.74). "
         "작목 대표 계수(딸기0.934·방울0.910·완숙0.940·참외0.712)는 같은 산식을 5농가에 적용한 평균. "
         "파랑 셀(설치배수)을 바꾸면 계수 자동 재계산.").font = _f(9, False, "595959")
ws7.row_dimensions[rfac+2].height = 42

# ══ Sheet8: 감가상각 (산식·live ⓐ/ⓑ) ═════════════════════════════════════
ws8 = wb.create_sheet("감가상각(산식)")
ws8.sheet_view.showGridLines = False
for col, w in zip("ABCDEF", [24, 16, 10, 12, 17, 17]): ws8.column_dimensions[col].width = w
ws8.merge_cells("A1:F1")
ws8["A1"] = "감가상각 — 산식과 계산 (실 자산 예시 · 살아있는 수식)"
ws8["A1"].font = _f(13, True, "FFFFFF"); ws8["A1"].fill = _fill(C_TITLE); ws8["A1"].alignment = CEN; ws8.row_dimensions[1].height = 26
ws8.merge_cells("A2:F2")
ws8["A2"] = "두 관례 →  ⓐ 조사표: 연감가 = 취득가 ÷ 내용연수   |   ⓑ 개선: 연감가 = 취득가 × (1 − 잔존율) ÷ 내용연수 (잔존율 10%)"
ws8["A2"].font = _f(9, False, "595959"); ws8["A2"].alignment = LEF; ws8.row_dimensions[2].height = 22
DEP_HDR = ["자산(예시)", "취득가액(원)", "잔존율", "내용연수(년)", "ⓐ 조사표 연감가(원)", "ⓑ 개선 연감가(원)"]
for j, h in enumerate(DEP_HDR):
    c = ws8.cell(4, j+1, h); c.font = _f(9, True, "FFFFFF"); c.fill = _fill(C_HDR); c.alignment = CEN; c.border = BORDER
DEP_ASSETS = [
    ("완숙 유리온실(김선환)", 3600000000, 0.10, 30), ("딸기 하우스A(이병권)", 800000000, 0.10, 15),
    ("방울 하우스A(진화)", 435529480, 0.10, 15), ("참외 작업장(유준상)", 100000000, 0.10, 15),
    ("복합환경제어기(예)", 8000000, 0.10, 8),
]
d0 = 5
for i, (name, acq, res, life) in enumerate(DEP_ASSETS):
    r = d0 + i
    ws8.cell(r, 1, name).font = _f(9); ws8.cell(r, 1).alignment = LEF; ws8.cell(r, 1).border = BORDER
    ws8.cell(r, 2, acq).font = _f(9, False, BLUE); ws8.cell(r, 2).alignment = RIG; ws8.cell(r, 2).border = BORDER; ws8.cell(r, 2).number_format = '#,##0'
    ws8.cell(r, 3, res).font = _f(9, False, BLUE); ws8.cell(r, 3).alignment = CEN; ws8.cell(r, 3).border = BORDER; ws8.cell(r, 3).number_format = '0%'
    ws8.cell(r, 4, life).font = _f(9, False, BLUE); ws8.cell(r, 4).alignment = CEN; ws8.cell(r, 4).border = BORDER
    ca = ws8.cell(r, 5, f"=ROUND(B{r}/D{r},0)"); ca.font = _f(9); ca.alignment = RIG; ca.border = BORDER; ca.number_format = '#,##0'
    cb = ws8.cell(r, 6, f"=ROUND(B{r}*(1-C{r})/D{r},0)"); cb.font = _f(9, True); cb.alignment = RIG; cb.border = BORDER; cb.number_format = '#,##0'
d_last = d0 + len(DEP_ASSETS) - 1
rds = d_last + 1
ws8.cell(rds, 1, "합계 (연 감가상각)").font = _f(10, True, C_TITLE); ws8.cell(rds, 1).alignment = LEF; ws8.cell(rds, 1).border = BORDER; ws8.cell(rds, 1).fill = _fill("E3EFE7")
for c in (2, 3, 4): ws8.cell(rds, c).border = BORDER; ws8.cell(rds, c).fill = _fill("E3EFE7")
for col in (5, 6):
    L = "E" if col == 5 else "F"
    cc = ws8.cell(rds, col, f"=SUM({L}{d0}:{L}{d_last})"); cc.font = _f(10, True, C_HDR); cc.alignment = RIG; cc.border = BORDER; cc.fill = _fill("E3EFE7"); cc.number_format = '#,##0'
ws8.merge_cells(start_row=rds+2, start_column=1, end_row=rds+2, end_column=6)
ws8.cell(rds+2, 1, "※ 파랑 셀(취득가·잔존율·내용연수)을 조사표 실입력값·견적서로 바꾸면 자동 재계산. "
         "완숙 유리온실 예: ⓐ 36억÷30 = 120,000,000, ⓑ 36억×0.9÷30 = 108,000,000. "
         "'작목별 CAPEX 실측'·'실 자산등록부'의 연감가는 ⓐ, 서비스 ERP는 ⓑ.").font = _f(9, False, "595959")
ws8.row_dimensions[rds+2].height = 42

# ══ Sheet0: 개요·핵심 산식 (표지 시트) ══════════════════════════════════════
ws0 = wb.create_sheet("개요·핵심 산식")
ws0.sheet_view.showGridLines = False
for col, w in zip("ABC", [30, 44, 30]): ws0.column_dimensions[col].width = w
ws0.merge_cells("A1:C1")
ws0["A1"] = "스마트팜 CAPEX/OPEX 체계화 — 개선된 소득조사표 종합 모델"
ws0["A1"].font = _f(14, True, "FFFFFF"); ws0["A1"].fill = _fill(C_TITLE); ws0["A1"].alignment = CEN; ws0.row_dimensions[1].height = 28
ws0.merge_cells("A2:C2")
ws0["A2"] = ("근거: 농진청 소득조사표 4작목·20농가(각 5). 현행(집계 상각비 2줄) → 개선(3계층 자산등록부·감가상각 수식·"
            "시설완비도 계수·OPEX 3대분류 연동). 취득가·내용연수·사양은 조사표 '농가' 시트 실입력, 업체만 견적 보강 필요.")
ws0["A2"].font = _f(9, False, "595959"); ws0["A2"].alignment = LEF; ws0.row_dimensions[2].height = 34
# 핵심 산식 2개
ws0.merge_cells("A4:C4"); ws0.cell(4, 1, "■ 핵심 산식").font = _f(11, True, C_TITLE)
FORMS = [
    ("시설 완비도 계수", "= 0.60 + Σ(시설별 가중치 × 설치배수)", "설치배수 ○=1·고장=0.5·×=0 · 최대 1.00 · [시설완비도 계수(산식)] 시트"),
    ("연 감가상각 ⓐ(조사표)", "= 취득가액 ÷ 내용연수", "조사표 실측 집계 방식 · [감가상각(산식)] 시트"),
    ("연 감가상각 ⓑ(개선)", "= 취득가액 × (1 − 잔존율) ÷ 내용연수", "잔존율 10% · 서비스 ERP 방식"),
    ("작목 계수", "= Σ(농가 계수) ÷ 농가수", "작목당 5농가 평균"),
    ("농가 총취득가", "= Σ(자산 취득가)", "'실 자산등록부' 합"),
    ("수리유지(OPEX)", "≈ 취득가 × 1~3% / 년", "경험식 · [OPEX-CAPEX 연동] 시트"),
]
for j, h in enumerate(["지표", "산식", "비고·위치"]):
    c = ws0.cell(5, j+1, h); c.font = _f(9, True, "FFFFFF"); c.fill = _fill(C_HDR); c.alignment = CEN; c.border = BORDER
for i, (name, form, memo) in enumerate(FORMS):
    r = 6 + i
    ws0.cell(r, 1, name).font = _f(9, True, C_TITLE); ws0.cell(r, 1).alignment = LEF; ws0.cell(r, 1).border = BORDER
    ws0.cell(r, 2, form).font = _f(9, True); ws0.cell(r, 2).alignment = LEF; ws0.cell(r, 2).border = BORDER
    ws0.cell(r, 3, memo).font = _f(8.5, False, "595959"); ws0.cell(r, 3).alignment = LEF; ws0.cell(r, 3).border = BORDER
    if i % 2:
        for c in range(1, 4): ws0.cell(r, c).fill = _fill(C_SOFT)
# 시트 안내
base0 = 6 + len(FORMS) + 1
ws0.merge_cells(start_row=base0, start_column=1, end_row=base0, end_column=3)
ws0.cell(base0, 1, "■ 시트 구성").font = _f(11, True, C_TITLE)
GUIDE = [
    "① 개요·핵심 산식 — 본 시트(요약·산식·시트 안내)",
    "② CAPEX 계층 등록부 — 3계층 자산 프레임(업체·사양·취득가·내용연수·잔존율·감가 수식)",
    "③ 시설완비도 계수(산식) — SUMPRODUCT 살아있는 계산",
    "④ 작목별 시설 구성 — 표본농가 ○/✕ 실데이터",
    "⑤ 감가상각(산식) — ⓐ/ⓑ 살아있는 계산",
    "⑥ 실 자산등록부(20농가) — 251개 자산 취득가·내용연수·사양 실추출",
    "⑦ 작목별 CAPEX 실측 — 작목별 총취득가·연감가 집계",
    "⑧ OPEX-CAPEX 연동 — 3대분류(재료비·경비·노무비) 매트릭스",
    "⑨ 내용연수 표준·요약 — 자산군별 표준 수명연한",
]
for i, g in enumerate(GUIDE):
    ws0.merge_cells(start_row=base0+1+i, start_column=1, end_row=base0+1+i, end_column=3)
    ws0.cell(base0+1+i, 1, g).font = _f(9)

# ── 시트 순서 재정리 ────────────────────────────────────────────────────────
ORDER = ["개요·핵심 산식", "CAPEX 계층 등록부", "시설완비도 계수(산식)", "작목별 시설 구성",
         "감가상각(산식)", "실 자산등록부(20농가)", "작목별 CAPEX 실측", "OPEX-CAPEX 연동", "내용연수 표준·요약"]
wb._sheets.sort(key=lambda s: ORDER.index(s.title) if s.title in ORDER else 99)
wb.active = 0

OUT.parent.mkdir(exist_ok=True)
wb.save(OUT)
print(f"저장: {OUT} · 시트 {len(wb.sheetnames)}: {wb.sheetnames}")
