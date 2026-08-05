# -*- coding: utf-8 -*-
"""소득조사표 개선 — 산식(수식) 워크북 (.xlsx, 살아있는 Excel 수식).
① 시설완비도 계수 산식  ② 감가상각 산식  ③ 개선 전략·산식 요약.
파랑=입력값(바꾸면 자동 재계산), 검정=수식. 근거: _build_facility_factor_dataset.py·capex_cost.py."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = Path(r"C:/smart_farm/out/소득조사표_개선_산식.xlsx")
FONT = "맑은 고딕"
GREEN="1F7A4D"; TITLE="123D2A"; HEAD="2F9A62"; SOFT="F1F6F1"; INK="171F1A"
BLUE="0000FF"  # 입력값(산업표준 색)
thin=Side(style="thin",color="CCCCCC"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
def f(sz=10,b=False,color=INK): return Font(name=FONT,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
LEF=Alignment(horizontal="left",vertical="center",wrap_text=True)
RIG=Alignment(horizontal="right",vertical="center")

def title(ws,txt,span,sz=13):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=span)
    c=ws.cell(1,1,txt); c.font=f(sz,True,"FFFFFF"); c.fill=fill(TITLE); c.alignment=CEN
    ws.row_dimensions[1].height=26
def note(ws,row,txt,span):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span)
    c=ws.cell(row,1,txt); c.font=f(9,False,"595959"); c.alignment=LEF

wb=Workbook()

# ══ Sheet1: 시설완비도 계수 산식 ═══════════════════════════════════════════
ws=wb.active; ws.title="① 시설완비도 계수 산식"
ws.sheet_view.showGridLines=False
for col,w in zip("ABCDEF",[22,11,11,13,13,11]): ws.column_dimensions[col].width=w
title(ws,"시설 완비도 계수 — 산식과 계산",6)
note(ws,2,"계수 = 0.60(온실구조 공통) + Σ(시설별 가중치 × 설치배수)   ·   설치배수: ○=1.0 · 고장=0.5 · ×(미설치)=0   ·   가중치 합 0.40 → 계수 최대 1.00",6)
ws.row_dimensions[2].height=28
# 헤더
hdr=["시설 항목","가중치","딸기(류창영)","방울(박경종)","완숙(김선환)","참외(강석구)"]
for j,h in enumerate(hdr):
    c=ws.cell(4,j+1,h); c.font=f(9,True,"FFFFFF"); c.fill=fill(HEAD); c.alignment=CEN; c.border=BORDER
# (시설, 가중치, 표본농가 설치배수 딸기/방울/완숙/참외)
ROWS=[
    ("온실구조체(공통)",0.60, 1,   1,   1,   1),
    ("일중천장",       0.03, 1,   1,   1,   1),
    ("이중천장",       0.03, 0,   0,   0,   0),
    ("측창(환기)",     0.07, 1,   1,   0,   1),
    ("천정 보온스크린", 0.07, 1,   1,   1,   0),
    ("측면 보온스크린", 0.06, 1,   0,   0,   0),
    ("차광 스크린",    0.06, 1,   1,   1,   0),
    ("관수·관비장치",  0.08, 1,   1,   0,   0.5),
]
r0=5
for i,(name,w,d,b,g,c_) in enumerate(ROWS):
    r=r0+i
    ws.cell(r,1,name).font=f(9,True if i==0 else False); ws.cell(r,1).alignment=LEF; ws.cell(r,1).border=BORDER
    ws.cell(r,2,w).font=f(9,False,BLUE); ws.cell(r,2).alignment=CEN; ws.cell(r,2).border=BORDER; ws.cell(r,2).number_format='0.00'
    for j,v in enumerate([d,b,g,c_]):
        cell=ws.cell(r,3+j,v); cell.font=f(9,False,BLUE); cell.alignment=CEN; cell.border=BORDER; cell.number_format='0.0'
        if i==0: cell.fill=fill(SOFT)
r_last=r0+len(ROWS)-1
# 가중치 합
r_sum=r_last+1
ws.cell(r_sum,1,"가중치 합").font=f(9,True,TITLE); ws.cell(r_sum,1).alignment=LEF; ws.cell(r_sum,1).border=BORDER; ws.cell(r_sum,1).fill=fill(SOFT)
ws.cell(r_sum,2,f"=SUM(B{r0}:B{r_last})").font=f(9,True); ws.cell(r_sum,2).alignment=CEN; ws.cell(r_sum,2).border=BORDER; ws.cell(r_sum,2).number_format='0.00'
# 계수 = SUMPRODUCT(가중치, 설치배수)
r_fac=r_sum+1
ws.cell(r_fac,1,"시설 완비도 계수").font=f(10,True,TITLE); ws.cell(r_fac,1).alignment=LEF; ws.cell(r_fac,1).border=BORDER; ws.cell(r_fac,1).fill=fill("E3EFE7")
ws.cell(r_fac,2,"").border=BORDER; ws.cell(r_fac,2).fill=fill("E3EFE7")
for j,col in enumerate("CDEF"):
    cell=ws.cell(r_fac,3+j,f"=SUMPRODUCT($B${r0}:$B${r_last},{col}{r0}:{col}{r_last})")
    cell.font=f(11,True,GREEN); cell.alignment=CEN; cell.border=BORDER; cell.fill=fill("E3EFE7"); cell.number_format='0.000'
# 참고: 표본1곳 vs 5농가평균
note(ws,r_fac+2,"※ 위 계수는 각 작목 표본농가 1곳에 산식을 적용한 값(딸기0.97·방울0.91·완숙0.76·참외0.74). "
     "보고서의 작목 계수(딸기0.934·방울0.910·완숙0.940·참외0.712)는 같은 산식을 5농가에 각각 적용한 평균이다. "
     "파랑 셀(설치배수)을 ○=1·고장=0.5·×=0으로 바꾸면 계수가 자동 재계산된다.",6)
ws.row_dimensions[r_fac+2].height=44

# ══ Sheet2: 감가상각 산식 ═══════════════════════════════════════════════════
ws2=wb.create_sheet("② 감가상각 산식")
ws2.sheet_view.showGridLines=False
for col,w in zip("ABCDEF",[24,16,10,11,16,20]): ws2.column_dimensions[col].width=w
title(ws2,"감가상각 — 산식과 계산 (실 자산 예시)",6)
note(ws2,2,"두 관례 병기 →  ⓐ 조사표 실측: 연감가 = 취득가 ÷ 내용연수 (잔존율 미적용)   |   ⓑ 개선 표준: 연감가 = 취득가 × (1 − 잔존율) ÷ 내용연수 (잔존율 10%)",6)
ws2.row_dimensions[2].height=22
h2=["자산(예시)","취득가액(원)","잔존율","내용연수(년)","ⓐ 조사표 연감가(원)","ⓑ 개선 연감가(원)"]
for j,h in enumerate(h2):
    c=ws2.cell(4,j+1,h); c.font=f(9,True,"FFFFFF"); c.fill=fill(HEAD); c.alignment=CEN; c.border=BORDER
ASSETS=[
    ("완숙 유리온실(김선환)", 3600000000, 0.10, 30),
    ("딸기 하우스A(이병권)",  800000000, 0.10, 15),
    ("방울 하우스A(진화)",    435529480, 0.10, 15),
    ("참외 작업장(유준상)",   100000000, 0.10, 15),
    ("복합환경제어기(예)",     8000000, 0.10, 8),
]
a0=5
for i,(name,acq,res,life) in enumerate(ASSETS):
    r=a0+i
    ws2.cell(r,1,name).font=f(9); ws2.cell(r,1).alignment=LEF; ws2.cell(r,1).border=BORDER
    ws2.cell(r,2,acq).font=f(9,False,BLUE); ws2.cell(r,2).alignment=RIG; ws2.cell(r,2).border=BORDER; ws2.cell(r,2).number_format='#,##0'
    ws2.cell(r,3,res).font=f(9,False,BLUE); ws2.cell(r,3).alignment=CEN; ws2.cell(r,3).border=BORDER; ws2.cell(r,3).number_format='0%'
    ws2.cell(r,4,life).font=f(9,False,BLUE); ws2.cell(r,4).alignment=CEN; ws2.cell(r,4).border=BORDER
    ca=ws2.cell(r,5,f"=ROUND(B{r}/D{r},0)"); ca.font=f(9); ca.alignment=RIG; ca.border=BORDER; ca.number_format='#,##0'
    cb=ws2.cell(r,6,f"=ROUND(B{r}*(1-C{r})/D{r},0)"); cb.font=f(9,True); cb.alignment=RIG; cb.border=BORDER; cb.number_format='#,##0'
a_last=a0+len(ASSETS)-1
rs=a_last+1
ws2.cell(rs,1,"합계 (연 감가상각)").font=f(10,True,TITLE); ws2.cell(rs,1).alignment=LEF; ws2.cell(rs,1).border=BORDER; ws2.cell(rs,1).fill=fill("E3EFE7")
for c in (2,3,4): ws2.cell(rs,c).border=BORDER; ws2.cell(rs,c).fill=fill("E3EFE7")
for col in (5,6):
    L="E" if col==5 else "F"
    cc=ws2.cell(rs,col,f"=SUM({L}{a0}:{L}{a_last})"); cc.font=f(10,True,GREEN); cc.alignment=RIG; cc.border=BORDER; cc.fill=fill("E3EFE7"); cc.number_format='#,##0'
note(ws2,rs+2,"※ 파랑 셀(취득가·잔존율·내용연수)을 조사표 '농가' 시트 실입력값 또는 견적서 값으로 바꾸면 자동 재계산. "
     "완숙 유리온실 예: ⓐ 36억÷30 = 120,000,000원/년, ⓑ 36억×0.9÷30 = 108,000,000원/년. "
     "보고서 제5장의 '평균 연감가'는 ⓐ(조사표 방식), 서비스 ERP·표준 프레임은 ⓑ(잔존율 적용)를 쓴다.",6)
ws2.row_dimensions[rs+2].height=46

# ══ Sheet3: 개선 전략·산식 요약 ════════════════════════════════════════════
ws3=wb.create_sheet("③ 개선 전략·산식")
ws3.sheet_view.showGridLines=False
for col,w in zip("ABC",[26,40,30]): ws3.column_dimensions[col].width=w
title(ws3,"개선 전략 — 방향과 산식 요약",3)
note(ws3,2,"현행(집계 상각비 2줄) → 개선(3계층 자산등록부·수식화·OPEX 연동). 아래는 각 단계의 방향과 적용 산식이다.",3)
h3=["단계(방향)","내용","적용 산식"]
for j,h in enumerate(h3):
    c=ws3.cell(4,j+1,h); c.font=f(9,True,"FFFFFF"); c.fill=fill(HEAD); c.alignment=CEN; c.border=BORDER
STRAT=[
    ("1. 자산 3계층 분해","상각비 2줄 → 대분류>중분류>세부품목, 자산별 업체·사양·취득가·내용연수·잔존율 부여","(구조화 — 등록부)"),
    ("2. 감가상각 수식화","자산별 정액 감가상각으로 매년 비용 산출·검증","연감가 = 취득가×(1−잔존율)÷내용연수"),
    ("3. 시설완비도 계수","작목·농가별 시설 완비 정도를 0~1 지표로 정량화","계수 = 0.60 + Σ(가중치×설치배수)"),
    ("4. 작목 계수 집계","작목 대표값 = 5농가 계수 평균","작목계수 = Σ(농가계수)÷농가수"),
    ("5. CAPEX 집계","농가 총투자비·평균 산출","총취득가 = Σ자산취득가 · 평균 = Σ농가÷농가수"),
    ("6. OPEX 연동","자산이 유발하는 운영비 매핑(에너지·수리유지·임차)","수리유지 ≈ 취득가 × 1~3%/년 (경험식)"),
    ("7. ERP·소득 정합","감가·수리유지를 비용모델·소득에 자동 반영","소득 = 매출 − (변동비 + 감가 + 수리유지)"),
]
s0=5
for i,(step,desc,formula) in enumerate(STRAT):
    r=s0+i
    ws3.cell(r,1,step).font=f(9,True,TITLE); ws3.cell(r,1).alignment=LEF; ws3.cell(r,1).border=BORDER
    ws3.cell(r,2,desc).font=f(9); ws3.cell(r,2).alignment=LEF; ws3.cell(r,2).border=BORDER
    cell=ws3.cell(r,3,formula); cell.font=f(9,True,GREEN); cell.alignment=LEF; cell.border=BORDER
    if i%2:
        for c in range(1,4): ws3.cell(r,c).fill=fill(SOFT)
note(ws3,s0+len(STRAT)+1,"※ 3·4는 '① 시설완비도 계수 산식' 시트, 2·5는 '② 감가상각 산식' 시트에서 살아있는 수식으로 계산된다. "
     "선행조건: 자산별 취득가·내용연수·사양은 조사표에 실입력(활용 가능), 업체(제조사)만 견적서 연동으로 보강 필요.",3)
ws3.row_dimensions[s0+len(STRAT)+1].height=40

OUT.parent.mkdir(exist_ok=True)
wb.save(OUT)
print(f"저장: {OUT} · 시트 {wb.sheetnames}")
