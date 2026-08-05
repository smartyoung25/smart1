# -*- coding: utf-8 -*-
"""소득조사표 개선 정리 — 발주처용 엑셀 (현황·문제점·개선점 + 품목별 실측 대조).
근거: 20농가 실측(작목당 5, 각 900㎡) · out/facility_factors.json.
값은 아티팩트/발주처 docx와 동일한 실측 기준."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(r"C:/smart_farm/out/소득조사표_개선_정리.xlsx")

# ── 팔레트 (아티팩트와 통일) ─────────────────────────────
GREEN   = "1F7A4D"; GREEN_SOFT = "E3EFE7"
CLAY    = "9C6A2E"; CLAY_SOFT  = "F2E8D6"
WARN    = "B0472E"
INK     = "171F1A"; INKSOFT = "54605A"
LINE    = "DBE2D7"; PAPER   = "F4F6F1"
WHITE   = "FFFFFF"

F = "Arial"
def font(sz=10, b=False, color=INK): return Font(name=F, size=sz, bold=b, color=color)
def fill(c): return PatternFill("solid", fgColor=c)
thin = Side(style="thin", color=LINE)
med  = Side(style="medium", color=GREEN)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
def align(h="left", v="center", wrap=True): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

TEMPLATE = 1_458_000  # 영농시설 783,000 + 대농구 675,000 (900㎡)

def style(ws, cell, val, sz=10, b=False, color=INK, bg=None, h="left", border=True):
    c = ws[cell]; c.value = val; c.font = font(sz, b, color); c.alignment = align(h)
    if bg: c.fill = fill(bg)
    if border: c.border = BORDER
    return c

def title_row(ws, row, text, span, bg=GREEN, color=WHITE, sz=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font(sz, True, color); c.fill = fill(bg); c.alignment = align("left")
    ws.row_dimensions[row].height = 26

# ══════════════════════════════════════════════════════════
wb = Workbook()

# ── 시트 1: 현황·문제점·개선점 (통합) ──────────────────────
ws = wb.active; ws.title = "현황·문제점·개선점"
ws.sheet_view.showGridLines = False
widths = [4, 34, 40, 24]
for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

r = 1
c = ws.cell(row=r, column=1, value="소득조사표 CAPEX/OPEX 체계화 — 현황·문제점·개선점")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c.font = font(15, True, GREEN); c.alignment = align("left"); ws.row_dimensions[r].height = 30
r += 1
c = ws.cell(row=r, column=1, value="근거: 농진청 소득조사표 4작목 · 20농가 실측(작목당 5) · 각 900㎡ · 연동 capex_cost→m4_cost→ERP /costs")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c.font = font(9, False, INKSOFT); c.alignment = align("left"); ws.row_dimensions[r].height = 18
r += 2

# 현황
title_row(ws, r, "Ⅰ. 현행 현황", 4, bg=CLAY); r += 1
style(ws, f"A{r}", "", border=False)
c = ws.cell(row=r, column=2, value="투자비(CAPEX)를 ‘상각비 2줄’로만 기록")
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
c.font = font(11, True, INK); c.alignment = align("left"); c.fill = fill(CLAY_SOFT); c.border = BORDER
ws.cell(row=r, column=1).fill = fill(CLAY_SOFT); ws.cell(row=r, column=1).border = BORDER
ws.row_dimensions[r].height = 22; r += 1
for txt in ["소득분석2 = 대농구 675,000 + 영농시설 783,000",
            "이름 ‘홍길동’·4작목 완전 동일 → 실 농가값이 아닌 샘플 템플릿",
            "실제 차이(생산량·시설유무)는 ‘농가’ 시트에 정성 코드(○/✕)로만 흩어짐"]:
    c = ws.cell(row=r, column=2, value="· " + txt)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    c.font = font(9.5, False, INKSOFT); c.alignment = align("left"); c.border = BORDER
    ws.cell(row=r, column=1).border = BORDER
    ws.row_dimensions[r].height = 17; r += 1
r += 1

# 문제점
title_row(ws, r, "Ⅱ. 문제점 — 조사표에서 실제 확인한 7대 현안", 4, bg=CLAY); r += 1
for h, col, w in [("#", 1, None), ("현안", 2, None), ("근거(조사표)", 3, None), ("영향", 4, None)]:
    style(ws, f"{get_column_letter(col)}{r}", h, b=True, color=WHITE, bg=CLAY, h="center" if col == 1 else "left")
ws.row_dimensions[r].height = 20; r += 1
PROBS = [
    ("1", "CAPEX가 집계 상각비 2줄뿐", "소득분석2 R21~23", "자산별 취득가·업체·성능·수명 추적 불가"),
    ("2", "샘플 템플릿이 실데이터처럼 존재", "‘홍길동’·4작목 675k/783k 동일", "실 투자비가 아님"),
    ("3", "분류가 2종(대농구·영농시설)뿐", "복합제어·양액기·센서 혼재", "스마트팜 자동화 설비 미반영"),
    ("4", "감가상각 산출근거 불투명", "취득가·내용연수·잔존율 부재", "재계산·검증·재투자 판단 불가"),
    ("5", "OPEX↔CAPEX 연동 단절", "수리유지·수도광열 대부분 0", "투자→운영비 파급 분석 불가"),
    ("6", "시설이 정성 코드로만", "‘농가’ 시설현황 ○/✕·취득가 결측", "정량 투자비 산출 불가"),
    ("7", "작목별 시설차이 미반영", "실측 완비도 0.71~0.94인데 상각비 4작목 동일", "작목 CAPEX 차이가 소득에 안 잡힘"),
]
for no, issue, ev, imp in PROBS:
    style(ws, f"A{r}", no, b=True, color=CLAY, h="center")
    style(ws, f"B{r}", issue, b=True)
    style(ws, f"C{r}", ev, sz=9, color=INKSOFT)
    style(ws, f"D{r}", imp, sz=9, color=INKSOFT)
    ws.row_dimensions[r].height = 30; r += 1
r += 1

# 개선점
title_row(ws, r, "Ⅲ. 개선점 — 3계층 자산등록부 + 연동 (구축 완료)", 4, bg=GREEN); r += 1
for h, col in [("개선", 2), ("내용", 3), ("해소 현안", 4)]:
    pass
style(ws, f"A{r}", "", bg=GREEN, border=True)
style(ws, f"B{r}", "개선", b=True, color=WHITE, bg=GREEN)
style(ws, f"C{r}", "내용", b=True, color=WHITE, bg=GREEN)
style(ws, f"D{r}", "해소 현안", b=True, color=WHITE, bg=GREEN, h="center")
ws.row_dimensions[r].height = 20; r += 1
IMPS = [
    ("3계층 자산등록부", "대분류(영농시설/기자재/대농기구) > 중분류 > 세부품목", "#1·#3"),
    ("자산별 11속성 + 정액감가상각 수식", "업체·성능·사양·취득가·내용연수·잔존율·감가상각", "#1·#4·#6"),
    ("표준 내용연수 명시", "법인세법 별표·농진청(온실15~20·스크린5~7·제어/관수7~10·센서5)", "#4"),
    ("OPEX↔CAPEX 연동 매트릭스", "수리유지→시설/대농기구 · 수도광열→제어·난방·관수", "#5"),
    ("작목별 시설구성 실데이터", "○설치 자산만 감가대상 · 작목 차등", "#7"),
    ("비용모델·ERP 자동연동", "register 취득가 → 감가상각·수리유지 자동 → 소득구조 반영", "#4·#5"),
    ("정직성 표기", "견적 입력 전엔 source=template로 명시", "#2"),
]
for name, desc, solve in IMPS:
    style(ws, f"A{r}", "✓", b=True, color=GREEN, h="center")
    style(ws, f"B{r}", name, b=True)
    style(ws, f"C{r}", desc, sz=9, color=INKSOFT)
    style(ws, f"D{r}", solve, b=True, color=GREEN, bg=GREEN_SOFT, h="center")
    ws.row_dimensions[r].height = 26; r += 1
r += 1
c = ws.cell(row=r, column=1, value="핵심 전환:  “영농시설 상각비 783,000”(한 줄)  →  자산별 분해 + 「취득가×(1−잔존율)÷수명」 수식")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c.font = font(10, True, GREEN); c.fill = fill(GREEN_SOFT); c.alignment = align("left"); c.border = BORDER
ws.row_dimensions[r].height = 24

# ── 시트 2: 품목별 실측 대조 ───────────────────────────────
ws2 = wb.create_sheet("품목별 실측 대조")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 24
for col in "BCDE": ws2.column_dimensions[col].width = 16

r = 1
c = ws2.cell(row=r, column=1, value="품목별 소득조사표 개선 — 20농가 실측 대조")
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c.font = font(15, True, GREEN); c.alignment = align("left"); ws2.row_dimensions[r].height = 30; r += 1
c = ws2.cell(row=r, column=1, value="계수·생산량·범위 = 작목별 5농가 실측 평균 · 이후 감가 = 템플릿 1,458,000 × 완비도계수 · 월감가 = 감가÷900㎡÷12")
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c.font = font(9, False, INKSOFT); c.alignment = align("left"); ws2.row_dimensions[r].height = 16; r += 2

# 헤더
crops = ["🍓 딸기", "🍅 방울토마토", "🍅 완숙토마토", "🍈 참외"]
style(ws2, f"A{r}", "항목", b=True, color=WHITE, bg=GREEN)
for i, cr in enumerate(crops):
    style(ws2, f"{get_column_letter(2+i)}{r}", cr, b=True, color=WHITE, bg=GREEN, h="center")
ws2.row_dimensions[r].height = 22; r += 1

# 데이터 행: (라벨, [딸기,방울,완숙,참외], 서식)  fmt: 'factor','range','int','money','money1','tmpl'
FACT = [0.934, 0.910, 0.940, 0.712]
PROD = [30448, 78344, 164314, 78768]
AFTER = [round(TEMPLATE * f) for f in FACT]
MONTH = [round(a / 900 / 12, 1) for a in AFTER]
ROWS = [
    ("시설 완비도 (5농가 평균)", FACT, "factor", True),
    ("완비도 범위 (농가 편차)", ["0.91–0.97", "0.76–0.97", "0.76–1.00", "0.67–0.75"], "text", False),
    ("평균 생산량 (kg)", PROD, "int", False),
    ("생산량 범위 (kg)", ["12k–59.5k", "58k–112k", "50k–450k", "39k–120k"], "text", False),
    ("이전 감가 (템플릿)", [TEMPLATE]*4, "int", False),
    ("이후 감가 (실측)", AFTER, "int", True),
    ("월 감가 (원/㎡)", MONTH, "money1", False),
]
LOW_IDX = 3  # 참외 열 강조
for label, vals, fmt, strong in ROWS:
    style(ws2, f"A{r}", label, b=strong, color=INK if strong else INKSOFT)
    for i, v in enumerate(vals):
        cell = f"{get_column_letter(2+i)}{r}"
        c = ws2[cell]; c.value = v; c.border = BORDER; c.alignment = align("center")
        is_low = (fmt in ("factor", "int", "money1") and i == LOW_IDX and strong)
        c.font = font(11 if strong else 10, strong, WARN if is_low else (INK if strong else INKSOFT))
        if fmt == "factor": c.number_format = "0.000"
        elif fmt == "int": c.number_format = "#,##0"
        elif fmt == "money1": c.number_format = "0.0"
        if strong: c.fill = fill(GREEN_SOFT if not is_low else "F6E2DB")
    ws2.row_dimensions[r].height = 22 if not strong else 24; r += 1
r += 1

# 품목별 요지
title_row(ws2, r, "품목별 요지 (문제 → 개선 → 활용)", 5, bg=GREEN); ws2_span = True; r += 1
STORIES = [
    ("🍓 딸기 (완비·노동집약)", "완비 시설이 템플릿에 안 잡힘 → 자산별 분해로 보온 투자·감가 명시 → 관리 자동화(유인·적엽) ROI 대조. 완비도 편차 작음(표준화)."),
    ("🍅 방울토마토 (다수확·측면보온 공백)", "측면보온 부재 미반영 → 표본 측면보온만 ✕로 공백 노출 → 측면보온 신규투자 시뮬레이션. 완비도 편차 최대(0.76~0.97)."),
    ("🍅 완숙토마토 (최다 생산·대부분 완비) ★정정", "표본은 측창·관수 ✕였으나 5농가 평균 0.940=대부분 완비. 자동화 공백은 전반이 아니라 농가 편차 → 시설 미비 개별 농가에 한해 자동화 ROI."),
    ("🍈 참외 (반촉성·저투자)", "딸기와 동일 투자비로 잡혀 소득 왜곡 → 감가 최소(−24%) 저CAPEX 실측 확인 → 관수 ‘고장’ 재투자 vs 수리 갈림길, 신규진입 저CAPEX 표준."),
]
for name, story in STORIES:
    c = ws2.cell(row=r, column=1, value=name)
    c.font = font(10, True, INK); c.alignment = align("left", wrap=False); c.border = BORDER
    c2 = ws2.cell(row=r, column=2, value=story)
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    c2.font = font(9, False, INKSOFT); c2.alignment = align("left"); c2.border = BORDER
    for cc in range(3, 6): ws2.cell(row=r, column=cc).border = BORDER
    ws2.row_dimensions[r].height = 40; r += 1

# ── 시트 3: 활용 로드맵 & 선행조건 ─────────────────────────
ws3 = wb.create_sheet("활용 로드맵·선행조건")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 4
ws3.column_dimensions["B"].width = 30
ws3.column_dimensions["C"].width = 62
r = 1
c = ws3.cell(row=r, column=1, value="활용 로드맵 & 선행조건")
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c.font = font(15, True, GREEN); c.alignment = align("left"); ws3.row_dimensions[r].height = 30; r += 2
title_row(ws3, r, "활용 로드맵", 3, bg=GREEN); r += 1
ROAD = [
    ("견적 연동 자동화", "C16 업로드 견적서·시방서에서 취득가·업체·사양 추출 → register 자동 채움"),
    ("투자 ROI 시뮬레이션", "신규 CAPEX(연 감가+유발 OPEX) vs 노무·수율·품질 편익"),
    ("재투자·교체 알림", "내용연수 만료·‘고장’ 자산을 재투자 vs 수리로 구분"),
    ("표준 CAPEX 벤치마크", "20개소 견적 축적 → 작목·규모·지역별 투자·감가 표준"),
    ("소득 모델 정합", "stage3·4 소득 검증에 실 감가·수리유지 반영(flat 고정비 대체)"),
    ("정책·보조사업 근거", "ROI·감가 구조 표준화로 심사·다년 손익분기 자료"),
]
for i, (t, d) in enumerate(ROAD, 1):
    style(ws3, f"A{r}", str(i), b=True, color=GREEN, h="center")
    style(ws3, f"B{r}", t, b=True)
    style(ws3, f"C{r}", d, sz=9, color=INKSOFT)
    ws3.row_dimensions[r].height = 24; r += 1
r += 1
title_row(ws3, r, "선행조건 (정직성)", 3, bg=CLAY); r += 1
c = ws3.cell(row=r, column=1,
    value="생산량·시설구성은 실데이터이나, 자산별 취득가(견적)·OPEX는 조사표에 미수집. "
          "현재 계수·생산량만 실측이고 감가상각 금액은 표준 템플릿 기반 — "
          "과업①에서 견적서·시방서를 조사항목에 포함하면 개선안이 실데이터로 즉시 작동.")
ws3.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=3)
c.font = font(10, False, INK); c.fill = fill(CLAY_SOFT); c.alignment = align("left"); c.border = BORDER
ws3.row_dimensions[r].height = 24

OUT.parent.mkdir(exist_ok=True)
wb.save(OUT)
print(f"저장: {OUT} ({OUT.stat().st_size:,} bytes) · 3시트")
