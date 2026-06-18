# -*- coding: utf-8 -*-
"""
KAASA smartfarmingsight — 기자재·시공업체 공식 참조 데이터 빌더
================================================================
스마트팜코리아 공식 DB(xlsx)를 읽어 공식 분류체계(스마트팜 ICT 기자재 표준)에
따라 재분류한 JSON 참조 데이터를 생성한다.

지속 업그레이드: 새 버전의 DB 엑셀이 나오면 --src 경로만 바꿔 재실행하면
api/data/reference/*.json 이 전량 갱신된다. (멱등)

사용:
  python scripts/build_equipment_reference.py \
    --src "C:/.../스마트팜코리아_DB(231025)_기자재정보_시공업체리스트_250117.xlsx"

산출:
  api/data/reference/equipment_taxonomy.json   공식 분류체계(대분류→표준장치명)
  api/data/reference/vendor_products.json      제조사 자사제품(표준장치명·모델·KC)
  api/data/reference/vendors.json              기업(제조/유통/시공) 디렉터리
  api/data/reference/construction_companies.json 온실 시공업체 도급순위(1~2군)
  api/data/reference/_meta.json                출처·버전·집계
"""
import argparse, json, os, re, sys
from collections import Counter, defaultdict
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(ROOT, "api", "data", "reference")

# ── 공식 분류체계 ────────────────────────────────────────────────
# 스마트팜 ICT 기자재 표준(농식품부/TTA) 3계층: 대분류 → 표준 장치명
# 스마트팜코리아 DB의 '표준 장치명'을 공식 대분류로 정규화 매핑한다.
TAXONOMY = {
    "감지부(센서)": {
        "desc": "생육·환경 계측 센서 노드",
        "devices": ["생체정보수집기", "축산용체중기"],
    },
    "제어부(구동기)": {
        "desc": "환경·구동 액추에이터 및 통합 환경제어",
        "devices": ["환경제어기", "환풍기", "냉방기", "안개분무시스템", "악취제거기"],
    },
    "양액·관수": {
        "desc": "양액 조성·관수 공급 장치",
        "devices": ["양액기", "자동급수기"],
    },
    "축산설비": {
        "desc": "축산 급이·방역·선별·냉각 설비",
        "devices": ["축산급이기", "사료빈관리기", "축산방역기",
                    "출하돈선별기", "원유냉각기"],
    },
    "기타·미분류": {
        "desc": "표준 장치명 미지정/기타",
        "devices": ["기타", ""],
    },
}
# 표준장치명 → 대분류 역인덱스
_DEV2CAT = {}
for cat, info in TAXONOMY.items():
    for d in info["devices"]:
        _DEV2CAT[d] = cat


def classify(std_name):
    s = (std_name or "").strip()
    return _DEV2CAT.get(s, "기타·미분류")


def _norm(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in ("None", "-", "nan") else s


def _rows(wb, sheet):
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    return hdr, rows[1:]


def _ci(hdr, name):
    """헤더 인덱스(없으면 -1)"""
    return hdr.index(name) if name in hdr else -1


# ── 1) 자사제품 → 분류 포함 제품 리스트 ─────────────────────────
def build_vendor_products(wb):
    hdr, rows = _rows(wb, "기업정보_자사제품")
    i_co, i_field = _ci(hdr, "기업명"), _ci(hdr, "활용 분야")
    i_std, i_maker = _ci(hdr, "표준 장치명"), _ci(hdr, "제조사")
    i_model = _ci(hdr, "모델명(제조/유통)")
    i_kc = _ci(hdr, "KC인증")
    out = []
    cat_count = Counter()
    for r in rows:
        std = _norm(r[i_std])
        if not _norm(r[i_co]) and not std:
            continue
        cat = classify(std)
        cat_count[cat] += 1
        out.append({
            "company": _norm(r[i_co]),
            "field": _norm(r[i_field]),            # 시설원예/축산
            "category": cat,                       # 공식 대분류
            "std_device": std,                     # 표준 장치명
            "maker": _norm(r[i_maker]),
            "model": _norm(r[i_model]),
            "kc": _norm(r[i_kc]),                  # 인증/미인증/인증제외
        })
    return out, cat_count


# ── 2) 기업 디렉터리(제조/유통/시공) ────────────────────────────
def build_vendors(wb):
    hdr, rows = _rows(wb, "기업정보")
    g = lambda r, n: _norm(r[_ci(hdr, n)]) if _ci(hdr, n) >= 0 else ""
    sectors = ["시설원예", "한우", "낙농", "육우", "양돈", "양계",
               "오리", "사슴", "양봉", "곤충", "말"]
    trade = ["제조(도매)", "제조(소매)", "수입(제조, 도매)", "수입(소매)",
             "국내(도매)", "국내(소매)", "정보통신공사 전문"]
    out = []
    for r in rows:
        name = g(r, "기업명")
        if not name:
            continue
        out.append({
            "name": name,
            "ceo": g(r, "대표자"),
            "founded": g(r, "설립년도"),
            "biz_type": g(r, "업종"),
            "biz_item": g(r, "업태"),
            "sectors": [s for s in sectors if g(r, s) == "O"],
            "trade": [t for t in trade if g(r, t) == "O"],
            "ict_products": g(r, "ICT 제품"),
            "address": g(r, "도로명 주소") or g(r, "지번주소"),
            "zipcode": g(r, "우편번호"),
            "tel": g(r, "대표번호"),
            "fax": g(r, "FAX번호"),
            "email": g(r, "E-MAIL"),
            "homepage": g(r, "홈페이지"),
            "as_staff": g(r, "AS가능인원"),
        })
    return out


# ── 3) 온실 시공업체 도급순위(참고자료 시트) ─────────────────────
_PHONE = re.compile(r"\d{2,4}-\d{3,4}-\d{4}|1\d{3}-\d{4}|\d{9,11}")


def build_construction(wb):
    out = []
    cur_rank = ""
    for sheet in ("참고자료", "Sheet1"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for r in ws.iter_rows(values_only=True):
            vals = [_norm(x) for x in r]
            cells = [v for v in vals if v]
            if not cells:
                continue
            # 군별 갱신(1군/2군 등)
            for v in cells:
                if re.fullmatch(r"[12]군", v):
                    cur_rank = v
            # 상호·주소·연락처 패턴: 마지막 셀이 전화면 채택
            phone = next((v for v in cells if _PHONE.fullmatch(v.replace(" ", ""))), "")
            if not phone:
                continue
            # 전화 앞 셀들 중 주소(시/도 포함)·상호 추출
            non_phone = [v for v in cells if v != phone and not re.fullmatch(r"\d+", v)
                         and not re.fullmatch(r"[12]군", v)]
            if not non_phone:
                continue
            addr = next((v for v in non_phone if re.search(
                r"(특별시|광역시|특별자치|도 |도$|시 |군 |구 )", v)), "")
            name = next((v for v in non_phone if v != addr), non_phone[0])
            out.append({
                "rank_group": cur_rank,
                "name": name,
                "address": addr,
                "tel": phone,
                "region": addr.split()[0] if addr else "",
            })
    # 중복 제거(상호+전화)
    seen, uniq = set(), []
    for c in out:
        k = (c["name"], c["tel"])
        if k in seen:
            continue
        seen.add(k)
        uniq.append(c)
    return uniq


def build_taxonomy(cat_count):
    cats = []
    for cat, info in TAXONOMY.items():
        cats.append({
            "category": cat,
            "desc": info["desc"],
            "std_devices": [d for d in info["devices"] if d],
            "product_count": cat_count.get(cat, 0),
        })
    return cats


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", required=True, help="스마트팜코리아 DB xlsx 경로")
    a = ap.parse_args()
    if not os.path.exists(a.src):
        sys.exit(f"[ERR] 원본 없음: {a.src}")
    wb = openpyxl.load_workbook(a.src, read_only=True, data_only=True)

    products, cat_count = build_vendor_products(wb)
    vendors = build_vendors(wb)
    construction = build_construction(wb)
    taxonomy = build_taxonomy(cat_count)

    os.makedirs(OUT_DIR, exist_ok=True)

    def dump(fn, obj):
        with open(os.path.join(OUT_DIR, fn), "w", encoding="utf-8") as f:
            json.dump(obj, f, ensure_ascii=False, indent=2)

    src_base = os.path.basename(a.src)
    ver = re.search(r"_(\d{6})\.xlsx$", src_base)
    meta = {
        "source": src_base,
        "source_version": ver.group(1) if ver else "",
        "official_standard": "스마트팜 ICT 기자재 표준(농식품부/TTA) — 표준 장치명 기준",
        "counts": {
            "vendor_products": len(products),
            "vendors": len(vendors),
            "construction_companies": len(construction),
            "taxonomy_categories": len(taxonomy),
        },
        "category_product_count": dict(cat_count),
    }
    dump("equipment_taxonomy.json", taxonomy)
    dump("vendor_products.json", products)
    dump("vendors.json", vendors)
    dump("construction_companies.json", construction)
    dump("_meta.json", meta)

    print("[OK] reference data built ->", OUT_DIR)
    print(json.dumps(meta, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
