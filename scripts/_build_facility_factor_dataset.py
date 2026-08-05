# -*- coding: utf-8 -*-
"""40개 조사표파일→20 고유농가 → 작목별 시설 완비도 계수 실측 재산정 + 생산량 데이터셋.

각 조사표 '농가' 시트에서 생산량·환경관리시설 설치유무·재배유형·난방원을 추출한다.
시설 완비도 계수 = base(온실구조 공통 0.60) + Σ(설치 시설 CAPEX 가중치, 고장=0.5).
작목별 계수 = 농가 평균. capex_cost 의 4샘플 추정 계수를 이 실측값으로 대체할 근거.

★ 정직성: 이 파일들은 생산량·시설구성은 실데이터(농가별 상이)이나, OPEX·자산별 취득가는
  비어 있다(조사표 양식 한계). 따라서 계수·생산량만 실측이고 감가상각 금액은 여전히 표준.

산출: out/소득조사표_시설계수_생산량_실측.xlsx + out/facility_factors.json
사용: PYTHONPATH=C:/smart_farm python scripts/_build_facility_factor_dataset.py
"""
from __future__ import annotations
import json
import glob
import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
BASE = Path(r"C:/hub/smartfarm-mvp/빅데이터생성형AI수익성향상")
OUT_XLSX = ROOT / "out" / "소득조사표_시설계수_생산량_실측.xlsx"
OUT_JSON = ROOT / "out" / "facility_factors.json"
FONT = "맑은 고딕"

# 폴더명 → 작목 표준명
CROP_DIR = {"딸기": "딸기", "방울토마토": "방울토마토", "참외": "참외", "토마토": "완숙토마토"}

# 농가 시트 환경관리시설 제어현황 행(B라벨) → (설치유무 열 I=9), CAPEX 가중치
FACILITIES = [
    ("일중천장", 86, 0.03), ("이중천장", 88, 0.03), ("측창", 90, 0.07),
    ("천정보온스크린", 92, 0.07), ("측면보온스크린", 94, 0.06),
    ("차광스크린", 96, 0.06), ("관수관비장치", 98, 0.08),
]
BASE_WEIGHT = 0.60   # 온실 구조체(공통) 가중치
PROD_ROW, PROD_COL = 120, 3      # 생산량
CULTIV_ROW, CULTIV_COL = 55, 10  # 재배유형 코드
HEAT_ROW, HEAT_COL = 69, 10      # 난방에너지원 코드


def _installed_value(v) -> float:
    """설치유무 셀 → 가중 배수: ○=1.0, 고장=0.5, 그 외/×=0."""
    if v is None:
        return 0.0
    s = str(v).strip()
    if "○" in s or s in ("O", "o", "1", "Y", "y"):
        return 1.0
    if "고장" in s:
        return 0.5
    return 0.0


def process_file(fp: str, crop: str) -> dict | None:
    try:
        wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
    except Exception as e:
        return {"error": str(e)[:40]}
    if "농가" not in wb.sheetnames:
        return None
    wg = wb["농가"]
    fac = {}
    factor = BASE_WEIGHT
    for name, row, weight in FACILITIES:
        raw = wg.cell(row, 9).value
        mult = _installed_value(raw)
        fac[name] = "○" if mult == 1.0 else ("고장" if mult == 0.5 else "×")
        factor += weight * mult
    prod = wg.cell(PROD_ROW, PROD_COL).value
    try:
        prod = float(prod) if prod not in (None, "") else None
    except Exception:
        prod = None
    return {
        "crop": crop,
        "farmer": Path(fp).stem.replace("스마트팜 경영데이터 조사표_", "").split("_")[0],
        "production_kg": prod,
        "cultivation_type": wg.cell(CULTIV_ROW, CULTIV_COL).value,
        "heating_code": wg.cell(HEAT_ROW, HEAT_COL).value,
        "factor": round(factor, 3),
        "facilities": fac,
    }


def main():
    rows = []
    for dirname, crop in CROP_DIR.items():
        for fp in glob.glob(str(BASE / "**" / f"*{crop if crop != '완숙토마토' else '완숙토마토'}*경영데이터 조사표*.xlsx"), recursive=True) \
                + glob.glob(str(BASE / "**" / dirname / "**" / "*경영데이터 조사표*.xlsx"), recursive=True):
            if "~$" in fp:
                continue
            rec = process_file(fp, crop)
            if rec and "error" not in rec and rec.get("factor"):
                rec["_fp"] = fp
                rows.append(rec)
    # 중복 제거 — 같은 농가의 버전 파일(검증/보완/v.2/v.3)을 농가 단위로 병합
    seen = {}; uniq = []
    for r in rows:
        key = (r["crop"], r["farmer"])
        if key in seen:
            continue
        seen[key] = True; uniq.append(r)
    n_files = len(rows)
    rows = uniq
    print(f"  파일 {n_files}건 → 고유 농가 {len(rows)}농가 (버전 중복 병합)")

    # 작목별 계수 집계
    by_crop = {}
    for r in rows:
        by_crop.setdefault(r["crop"], []).append(r)
    factors = {}
    for crop, recs in by_crop.items():
        facs = [x["factor"] for x in recs]
        prods = [x["production_kg"] for x in recs if x["production_kg"]]
        factors[crop] = {
            "n_farms": len(recs),
            "factor_mean": round(sum(facs) / len(facs), 3),
            "factor_min": round(min(facs), 3),
            "factor_max": round(max(facs), 3),
            "production_mean_kg": round(sum(prods) / len(prods), 0) if prods else None,
            "production_min_kg": round(min(prods), 0) if prods else None,
            "production_max_kg": round(max(prods), 0) if prods else None,
        }

    OUT_JSON.parent.mkdir(exist_ok=True)
    OUT_JSON.write_text(json.dumps({
        "_note": "40개 조사표파일→20 고유농가 '농가' 시트 실데이터. factor=시설완비도(base0.60+설치시설 가중), 고장=0.5. OPEX·취득가는 조사표에 미수집(계수·생산량만 실측).",
        "crop_factors": factors,
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    # ── xlsx ──
    from openpyxl import Workbook
    C_T, C_H, C_S = "1C5A3A", "2F9A62", "F4F7F4"
    thin = Side(style="thin", color="CCCCCC"); BD = Border(left=thin, right=thin, top=thin, bottom=thin)
    def F(sz=10, b=False, c="16211B"): return Font(name=FONT, size=sz, bold=b, color=c)
    wb = Workbook()

    # Sheet1: 작목별 계수 요약
    ws = wb.active; ws.title = "작목별 시설계수(실측)"
    ws.merge_cells("A1:H1"); ws["A1"] = "작목별 시설 완비도 계수 — 20농가(40파일 버전병합) 실측 재산정"
    ws["A1"].font = F(13, True, "FFFFFF"); ws["A1"].fill = PatternFill("solid", fgColor=C_T); ws["A1"].alignment = Alignment(horizontal="center")
    hdr = ["작목", "농가수", "계수(평균)", "계수(최소~최대)", "생산량평균(kg)", "생산량(최소~최대)", "종전(4샘플)추정", "차이"]
    prev = {"딸기": 1.00, "방울토마토": 0.92, "완숙토마토": 0.82, "참외": 0.70}
    widths = [14, 8, 11, 15, 14, 18, 14, 8]
    for i, w in enumerate(widths): ws.column_dimensions[get_column_letter(i+1)].width = w
    for i, h in enumerate(hdr):
        c = ws.cell(2, i+1, h); c.font = F(9, True, "FFFFFF"); c.fill = PatternFill("solid", fgColor=C_H); c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = BD
    for ri, (crop, s) in enumerate(sorted(factors.items(), key=lambda x: -x[1]["factor_mean"])):
        pv = prev.get(crop)
        vals = [crop, s["n_farms"], s["factor_mean"], f"{s['factor_min']}~{s['factor_max']}",
                f"{s['production_mean_kg']:,.0f}" if s["production_mean_kg"] else "-",
                f"{s['production_min_kg']:,.0f}~{s['production_max_kg']:,.0f}" if s["production_min_kg"] else "-",
                pv if pv else "-", (round(s["factor_mean"]-pv, 3) if pv else "-")]
        for ci, v in enumerate(vals):
            c = ws.cell(3+ri, ci+1, v); c.border = BD; c.font = F(9); c.alignment = Alignment(horizontal="center" if ci else "left")
            if ri % 2: c.fill = PatternFill("solid", fgColor=C_S)

    # Sheet2: 농가별 데이터셋
    ws2 = wb.create_sheet("농가별 데이터셋")
    cols = ["작목", "농가", "생산량(kg)", "재배유형", "난방코드", "시설계수"] + [f[0] for f in FACILITIES]
    for i in range(len(cols)):
        ws2.column_dimensions[get_column_letter(i+1)].width = 14 if i < 6 else 12
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    ws2["A1"] = f"농가별 생산량·시설 데이터셋 (n={len(rows)})"
    ws2["A1"].font = F(12, True, "FFFFFF"); ws2["A1"].fill = PatternFill("solid", fgColor=C_T); ws2["A1"].alignment = Alignment(horizontal="center")
    for i, h in enumerate(cols):
        c = ws2.cell(2, i+1, h); c.font = F(9, True, "FFFFFF"); c.fill = PatternFill("solid", fgColor=C_H); c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = BD
    for ri, r in enumerate(sorted(rows, key=lambda x: (x["crop"], -(x["production_kg"] or 0)))):
        base = [r["crop"], r["farmer"], (round(r["production_kg"]) if r["production_kg"] else "-"),
                r["cultivation_type"], r["heating_code"], r["factor"]] + [r["facilities"][f[0]] for f in FACILITIES]
        for ci, v in enumerate(base):
            c = ws2.cell(3+ri, ci+1, v); c.border = BD; c.font = F(9); c.alignment = Alignment(horizontal="center" if ci != 1 else "left")
            if ci == 2 and isinstance(v, int): c.number_format = '#,##0'
    ws2.freeze_panes = "A3"

    OUT_XLSX.parent.mkdir(exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"  저장: {OUT_XLSX}")
    print(f"  저장: {OUT_JSON}")
    print("\n  ── 작목별 계수(실측) vs 종전(4샘플) ──")
    for crop, s in sorted(factors.items(), key=lambda x: -x[1]["factor_mean"]):
        pv = prev.get(crop, "-")
        print(f"    {crop:<7} n={s['n_farms']:>2} 계수 {s['factor_mean']:.3f} (범위 {s['factor_min']}~{s['factor_max']}) · 종전 {pv} · 생산량 평균 {s['production_mean_kg']:,.0f}kg" if s['production_mean_kg'] else f"    {crop}: n={s['n_farms']}")


if __name__ == "__main__":
    main()
